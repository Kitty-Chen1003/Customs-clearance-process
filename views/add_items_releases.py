import os

import pandas as pd
from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtGui import QFont
from PyQt5.QtWidgets import QWidget, QVBoxLayout, QLabel, QLineEdit, QSizePolicy, QPushButton, QHBoxLayout, QFileDialog, \
    QMessageBox, QScrollArea, QApplication, QDialog, QSpacerItem, QListWidget, QAbstractItemView, QFrame, \
    QListWidgetItem
from openpyxl.reader.excel import load_workbook

import global_state
from http_client import add_awb_request, add_items_releases_request
from utils.parse_zc429_xml import parse_zc429_xml, parse_zc429_json


# === 模式选择对话框 ===
class ReleaseModeDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Select Processing Mode")
        self.setMinimumWidth(350)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(40, 30, 40, 30)  # 添加左右和上下内边距
        layout.setSpacing(20)

        label = QLabel("Please choose the processing mode:")
        label.setObjectName("normal")
        label.setAlignment(Qt.AlignCenter)
        layout.addWidget(label)

        self.more_than_150_btn = QPushButton(">150")
        self.more_than_150_btn.setObjectName("confirmbutton")
        self.more_than_150_btn.clicked.connect(lambda: self.accept_with_mode(">150"))
        layout.addWidget(self.more_than_150_btn)

        self.less_than_equal_150_btn = QPushButton("<=150")
        self.less_than_equal_150_btn.setObjectName("confirmbutton")
        self.less_than_equal_150_btn.clicked.connect(lambda: self.accept_with_mode("<=150"))
        layout.addWidget(self.less_than_equal_150_btn)

        layout.addItem(QSpacerItem(20, 10))

        self.selected_mode = None

    def accept_with_mode(self, mode):
        self.selected_mode = mode
        self.accept()


# === 多文件上传弹窗（用于 >150 和 手动模式） ===
class MultiFileUploadDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Upload ZC429 Files(>150)")
        self.setMinimumWidth(600)

        self.zc429_files = []

        layout = QVBoxLayout(self)
        layout.setContentsMargins(40, 30, 40, 30)
        layout.setSpacing(20)

        self.zc429_list = QListWidget()
        btn_429 = QPushButton("Upload ZC429 Files")
        btn_429.setFont(QFont("Microsoft YaHei", 12))
        btn_429.setMaximumWidth(800)
        btn_429.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        btn_429.setStyleSheet("padding: 10px;")
        btn_429.clicked.connect(lambda: self.upload_file(self.zc429_list, self.zc429_files))

        self.get_mrn_btn = QPushButton("Get Releases MRN")
        self.get_mrn_btn.setObjectName("confirmbutton")
        self.get_mrn_btn.clicked.connect(self.get_mrn)

        lab_429 = QLabel("ZC429 Files:")
        lab_429.setObjectName("normal")

        layout.addWidget(lab_429)
        layout.addWidget(self.zc429_list)
        layout.addWidget(btn_429)

        layout.addWidget(self.get_mrn_btn)

    def upload_file(self, list_widget, storage_list):
        files, _ = QFileDialog.getOpenFileNames(self, "Select Files", "", "XML Files (*.xml)")
        for f in files:
            if f not in storage_list:
                storage_list.append(f)
                list_widget.addItem(f)

    def get_mrn(self):
        QMessageBox.information(self, "Get Releases MRN", "Not Support yet.")
        self.accept()


# === 多文件上传弹窗（用于 >150 和 手动模式） ===
class OldmultiFileUploadDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)

        self.zc_429_file_path = None

        self.setWindowTitle("Upload ZC429 Files(<=150)")
        self.setMinimumWidth(600)

        self.zc429_files = []

        layout = QVBoxLayout(self)
        layout.setContentsMargins(40, 30, 40, 30)
        layout.setSpacing(20)

        self.zc429_list = QListWidget()
        btn_429 = QPushButton("Upload ZC429 Files")
        btn_429.setFont(QFont("Microsoft YaHei", 12))
        btn_429.setMaximumWidth(800)
        btn_429.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        btn_429.setStyleSheet("padding: 10px;")
        btn_429.clicked.connect(lambda: self.upload_file(self.zc429_list, self.zc429_files))

        self.get_mrn_btn = QPushButton("Get Releases MRN")
        self.get_mrn_btn.setObjectName("confirmbutton")
        self.get_mrn_btn.clicked.connect(self.get_mrn)

        lab_429 = QLabel("ZC429 Files:")
        lab_429.setObjectName("normal")

        layout.addWidget(lab_429)
        layout.addWidget(self.zc429_list)
        layout.addWidget(btn_429)

        layout.addWidget(self.get_mrn_btn)

    def upload_file(self, list_widget, storage_list):
        files, _ = QFileDialog.getOpenFileNames(self, "Select Files", "", "XML Files (*.xml)")
        for f in files:
            if f not in storage_list:
                storage_list.append(f)
                list_widget.addItem(f)

    def get_mrn(self):
        # 汇总所有 ZC429 文件中的数据
        all_release_info = []
        for file_path in self.zc429_files:
            try:
                with open(file_path, 'rb') as f:
                    all_release_info.extend(parse_zc429_xml(f))
                # with open(file_path, 'r', encoding='utf-8') as f:
                #     all_release_info.extend(parse_zc429_json(f))
            except Exception as e:
                print(f"Error parsing {file_path}: {e}")

        if not all_release_info:
            QMessageBox.warning(self, "Warning", "No valid MRN found in the uploaded files.")
            return

        # 打开并补全 Excel 文件
        if not global_state.process_file_path or not os.path.exists(global_state.process_file_path):
            QMessageBox.warning(self, "Error", "Excel file path not set or not exist.")
            return

        try:
            df = pd.read_excel(global_state.process_file_path)

            # 构建一个 dict 加快查找速度
            ucr_to_mrn = dict(all_release_info)

            # 写入 MRN(RELEASES) 列
            df["MRN(RELEASES)"] = df["TrackingNumber"].apply(lambda x: ucr_to_mrn.get(str(x).strip(), ""))

            # 覆盖写回原文件
            df.to_excel(global_state.process_file_path, index=False)

            QMessageBox.information(self, "Success", f"MRN写入成功，共匹配 {len(ucr_to_mrn)} 条。")
            self.accept()

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to update Excel file:\n{e}")


    # def get_mrn(self):
    #     # 汇总所有 ZC429 文件中的数据
    #     all_release_info = []
    #
    #     for file_path in self.zc429_files:
    #         try:
    #             with open(file_path, 'r', encoding='utf-8') as f:
    #                 all_release_info.extend(parse_zc429_json(f))
    #         except Exception as e:
    #             print(f"Error parsing {file_path}: {e}")
    #
    #     if not all_release_info:
    #         QMessageBox.warning(self, "Warning", "No valid MRN found in the uploaded files.")
    #         return
    #
    #     # ✅ 检查 Excel 是否存在
    #     if not global_state.process_file_path or not os.path.exists(global_state.process_file_path):
    #         QMessageBox.warning(self, "Error", "Excel file path not set or not exist.")
    #         return
    #
    #     try:
    #         # ✅ 构建 UCR → MRN 映射
    #         ucr_to_mrn = dict(all_release_info)
    #
    #         # ✅ 用 openpyxl 打开原 Excel（不破坏样式）
    #         wb = load_workbook(global_state.process_file_path)
    #         ws = wb.active  # 默认第一个 sheet
    #
    #         # ✅ 获取表头所在列
    #         header_row = 1
    #         headers = {}
    #         for col in range(1, ws.max_column + 1):
    #             val = ws.cell(row=header_row, column=col).value
    #             if val:
    #                 headers[val] = col
    #
    #         if "TrackingNumber" not in headers:
    #             QMessageBox.critical(self, "Error", "Excel 中未找到 TrackingNumber 列")
    #             return
    #
    #         # ✅ 如果 MRN(RELEASES) 列不存在，则创建（但不改其它样式）
    #         if "MRN(RELEASES)" not in headers:
    #             mrn_col = ws.max_column + 1
    #             ws.cell(row=1, column=mrn_col).value = "MRN(RELEASES)"
    #         else:
    #             mrn_col = headers["MRN(RELEASES)"]
    #
    #         tracking_col = headers["TrackingNumber"]
    #
    #         match_count = 0
    #
    #         # ✅ 从第二行开始逐行回填（不影响样式）
    #         for row in range(2, ws.max_row + 1):
    #             tracking_value = ws.cell(row=row, column=tracking_col).value
    #             if tracking_value:
    #                 tracking_value = str(tracking_value).strip()
    #                 mrn_value = ucr_to_mrn.get(tracking_value, "")
    #                 if mrn_value:
    #                     ws.cell(row=row, column=mrn_col).value = mrn_value
    #                     match_count += 1
    #
    #         # ✅ 只保存数据，不破坏样式
    #         wb.save(global_state.process_file_path)
    #
    #         QMessageBox.information(
    #             self,
    #             "Success",
    #             f"MRN写入成功，共匹配 {match_count} 条。"
    #         )
    #
    #         self.accept()
    #
    #     except Exception as e:
    #         QMessageBox.critical(self, "Error", f"Failed to update Excel file:\n{e}")
class AddItemsReleases(QWidget):
    # finished_successful = pyqtSignal()
    cancel_requested = pyqtSignal()

    def __init__(self, api):

        super().__init__()

        self.column_input = None
        self.step_config = None
        self.api = api
        # self.finished = False
        self.selected_files = []

        self.release_data = []  # [{ process_file: str, zc429_files: []}]

        self.outer_layout = None
        self.scroll_area = None
        self.inner_widget = None
        self.inner_layout = None
        self.scroll_area_label = None
        self.label_status = None

        self.process_list = None
        self.zc429_list = None

        self.label_mapping = None
        self.scroll_area_mapping = None

        self.suffix_input = None

        self.init_ui()

    def set_step_config(self, step_config: dict):
        """设置字段名和提示"""
        self.step_config = step_config
        self.update_hint_label()

    def update_hint_label(self):
        """在页面标题下方增加提示"""
        if not self.step_config or not hasattr(self, "inner_layout"):
            return

        # 删除已有提示标签（如果存在）
        for i in reversed(range(self.inner_layout.count())):
            widget = self.inner_layout.itemAt(i).widget()
            if isinstance(widget, QLabel) and getattr(widget, "is_hint_label", False):
                self.inner_layout.removeWidget(widget)
                widget.deleteLater()

        hint_text = self.step_config.get("hint", "")
        if hint_text:
            hint_label = QLabel(hint_text)
            hint_label.setObjectName("hint")
            hint_label.setAlignment(Qt.AlignCenter)
            hint_label.is_hint_label = True  # 标记为提示
            # 找到标题位置后插入到其下方
            # 假设标题总是 inner_layout 的第一个 QLabel
            for i in range(self.inner_layout.count()):
                widget = self.inner_layout.itemAt(i).widget()
                if isinstance(widget, QLabel) and widget.objectName() == "title":
                    self.inner_layout.insertWidget(i + 1, hint_label)
                    break
                    
    def init_ui(self):

        self.build_outer_layout()

        self.build_scroll_area()

        self.build_inner_layout()

        self.build_title()
        
        self.build_column_input()

        self.build_process_clearance_panel()

        self.build_action_buttons()

    # def select_file(self):
    #     file_dialog = QFileDialog(self)
    #     file_dialog.setWindowTitle('Select File')
    #     file_dialog.setFileMode(QFileDialog.ExistingFiles)  # 只允许选择一个文件
    #     file_dialog.setNameFilter("Supported Files (*.xml *.xls *.xlsx)")  # 限制文件类型
    #     file_dialog.setViewMode(QFileDialog.List)  # 以列表模式显示文件
    #
    #     if file_dialog.exec_():
    #         # 更新 self.selected_files 列表为用户选择的文件路径
    #         self.selected_files = file_dialog.selectedFiles()
    #         self.label_status.setText("\n".join(self.selected_files))  # 每个文件路径换行显示

    def build_outer_layout(self):

        # 外层布局（用于居中）
        self.outer_layout = QVBoxLayout(self)
        self.outer_layout.setContentsMargins(0, 0, 0, 0)
        self.outer_layout.setAlignment(Qt.AlignCenter)

    def build_scroll_area(self):
        # 创建 QScrollArea，并设置属性
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)  # 让 scroll area 根据内容自适应
        # self.scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)  # 横向不要滚动条（可选）
        self.scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)  # 纵向滚动条按需出现

        self.outer_layout.addWidget(self.scroll_area)

    def build_inner_layout(self):
        # 内层容器和布局（承载内容）
        self.inner_widget = QWidget()
        self.inner_layout = QVBoxLayout(self.inner_widget)
        self.inner_layout.setContentsMargins(100, 50, 100, 50)
        self.inner_layout.setSpacing(20)
        self.inner_layout.setAlignment(Qt.AlignTop | Qt.AlignHCenter)  # 控制内容居中靠上

        self.scroll_area.setWidget(self.inner_widget)

    def build_title(self):
        # 标题
        title = QLabel("AddItemsReleases")
        title.setObjectName("title")
        title.setAlignment(Qt.AlignCenter)
        self.inner_layout.addWidget(title)
        # self.inner_layout.addSpacing(50)

    def build_column_input(self):
        column_layout = QVBoxLayout()

        column_label = QLabel("Column Number:")
        column_label.setObjectName("normal")
        self.column_input = QLineEdit()
        self.column_input.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        column_layout.addWidget(column_label)
        column_layout.addWidget(self.column_input)
        column_layout.setAlignment(Qt.AlignHCenter)  # 整个控件居中

        self.inner_layout.addLayout(column_layout)
        self.inner_layout.addSpacing(10)
        
    def build_process_clearance_panel(self):

        process_release_layout = QHBoxLayout()

        # 左侧 arn 列表
        left_layout = QVBoxLayout()
        process_label = QLabel("Process Files:")
        process_label.setObjectName("normal")

        self.process_list = QListWidget()
        self.process_list.setSelectionMode(QAbstractItemView.SingleSelection)

        self.process_list.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.process_list.setMinimumHeight(300)

        # self.process_list.setSelectionBehavior(QAbstractItemView.SelectRows)
        # self.process_list.setEditTriggers(QAbstractItemView.NoEditTriggers)
        # self.process_list.horizontalHeader().setStretchLastSection(True)
        # self.process_list.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # self.process_list.verticalHeader().setVisible(False)
        # self.process_list.setSelectionMode(QAbstractItemView.SingleSelection)  # 单选模式

        # self.process_list.setStyleSheet("""
        #     QTableWidget {
        #         background-color: #ffffff;
        #         gridline-color: #ccc;
        #         border: 1px solid #aaa;
        #     }
        #     QHeaderView::section {
        #         background-color: #0078d7;
        #         color: white;
        #         padding: 6px 12px;
        #         border: 1px solid #aaa;
        #     }
        #     QTableWidget::item:selected {
        #         background-color: #87cefa;
        #         color: black;
        #     }
        # """)

        self.process_list.setStyleSheet("""
            QListWidget {
                background-color: #f9f9f9;
                border: 1px solid #ccc;
                padding: 5px;
            }
            QListWidget::item {
                padding: 5px;
            }
            QListWidget::item:selected {
                background-color: #87cefa;
                color: black;
            }
        """)

        self.process_list.itemClicked.connect(self.show_mapping)

        # TODO:按钮设置样式
        process_btn_layout = QHBoxLayout()

        select_all_btn = QPushButton("Select All / Unselect All")
        select_all_btn.setObjectName("fcAddButton")
        select_all_btn.clicked.connect(self.toggle_selects)

        add_process_btn = QPushButton("Add Process File")
        add_process_btn.setObjectName("fcAddButton")
        add_process_btn.clicked.connect(self.add_process_file)

        del_process_btn = QPushButton("Delete Process File")
        del_process_btn.setObjectName("fcDelButton")
        del_process_btn.clicked.connect(self.delete_process_file)

        process_btn_layout.addWidget(select_all_btn)
        process_btn_layout.addWidget(add_process_btn)
        process_btn_layout.addWidget(del_process_btn)

        left_layout.addWidget(process_label)
        left_layout.addWidget(self.process_list)
        left_layout.addLayout(process_btn_layout)

        # 右侧 mapping 文件列表
        right_layout = QVBoxLayout()

        mapping_label = QLabel("ZC429 Files:")
        mapping_label.setObjectName("normal")

        self.zc429_list = QListWidget()
        self.zc429_list.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.zc429_list.setMinimumHeight(300)
        self.zc429_list.setSelectionMode(QAbstractItemView.NoSelection)  # 单选模式

        self.zc429_list.setStyleSheet("""
            QListWidget {
                background-color: #f9f9f9;
                border: 1px solid #ccc;
                padding: 5px;
            }
            QListWidget::item {
                padding: 5px;
            }
            QListWidget::item:selected {
                background-color: #87cefa;
                color: black;
            }
        """)

        # mapping 操作按钮
        map_btn_layout = QHBoxLayout()
        add_429_btn = QPushButton("Add ZC429 Files")
        add_429_btn.clicked.connect(self.add_429_file)

        del_429_btn = QPushButton("Delete ZC429 Files")
        del_429_btn.clicked.connect(self.delete_429_file)

        add_429_btn.setObjectName("fcAddButton")
        del_429_btn.setObjectName("fcDelButton")

        map_btn_layout.addWidget(add_429_btn)
        map_btn_layout.addWidget(del_429_btn)

        right_layout.addWidget(mapping_label)
        right_layout.addWidget(self.zc429_list)
        right_layout.addLayout(map_btn_layout)

        process_429_group = QFrame()
        process_429_group.setObjectName("fcFrame")
        process_429_group.setFrameShape(QFrame.NoFrame)
        process_429_group.setLayout(process_release_layout)

        # fc_manifest_layout.addStretch()
        process_release_layout.addLayout(left_layout)
        process_release_layout.addSpacing(50)
        process_release_layout.addLayout(right_layout)
        # fc_manifest_layout.addStretch()

        process_release_layout.setStretchFactor(left_layout, 1)
        process_release_layout.setStretchFactor(right_layout, 1)

        # self.inner_layout.addLayout(fc_manifest_layout)
        self.inner_layout.addWidget(process_429_group)

        self.inner_layout.addSpacing(20)

    def build_action_buttons(self):
        suffix_layout = QVBoxLayout()

        suffix_label = QLabel("Tracking Number Suffix (e.g. A):")
        suffix_label.setObjectName("normal")
        self.suffix_input = QLineEdit()
        self.suffix_input.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        suffix_layout.addWidget(suffix_label)
        suffix_layout.addWidget(self.suffix_input)

        self.inner_layout.addLayout(suffix_layout)
        self.inner_layout.addSpacing(20)

        write_btn = QPushButton("Write Release to Process File")
        write_btn.setFont(QFont("Microsoft YaHei", 12))
        write_btn.setStyleSheet("padding: 10px;")
        write_btn.clicked.connect(self.write_release)

        self.inner_layout.addWidget(write_btn)
        self.inner_layout.addSpacing(20)

        self.get_mrn_button = QPushButton("Get Release MRN")
        self.get_mrn_button.setFont(QFont("Microsoft YaHei", 12))
        self.get_mrn_button.setStyleSheet("padding: 10px;")
        self.get_mrn_button.clicked.connect(self.get_release_mrn)

        self.inner_layout.addWidget(self.get_mrn_button)
        self.inner_layout.addSpacing(20)

        # 按钮水平布局（发送按钮和取消按钮并排）
        button_layout = QHBoxLayout()

        # 发送按钮
        submit_button = QPushButton("Submit")
        submit_button.setObjectName("confirmbutton")
        submit_button.clicked.connect(self.submit)
        # submit_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        button_layout.addWidget(submit_button)

        # 取消按钮
        cancel_button = QPushButton("Cancel")
        cancel_button.setObjectName("cancelButton")
        cancel_button.clicked.connect(self.on_cancel)  # 连接到提示框槽函数
        # cancel_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        button_layout.addWidget(cancel_button)

        # 将按钮布局添加到主布局中
        self.inner_layout.addLayout(button_layout)
        self.inner_layout.addSpacing(20)
        self.inner_layout.addStretch()

        self.save_button = QPushButton("Save_Release")
        self.save_button.setFont(QFont("Microsoft YaHei", 12))
        self.save_button.setStyleSheet("padding: 10px;")
        self.save_button.clicked.connect(self.save_release)
        self.save_button.setEnabled(False)

        self.inner_layout.addWidget(self.save_button)
        self.inner_layout.addSpacing(20)
        self.inner_layout.addStretch()

    def show_mapping(self, item):
        self.zc429_list.clear()

        row = self.process_list.row(item)  # ✅ QTableWidgetItem → int
        if row < 0 or row >= len(self.release_data):
            return

        data = self.release_data[row]

        # 显示 415
        for f in data.get("zc429_files", []):
            list_item = QListWidgetItem(f)
            list_item.setToolTip(f)
            self.zc429_list.addItem(list_item)

    def toggle_selects(self):
        if self.process_list.count() == 0:
            return

        # 判断是否已经全部选中
        all_checked = True
        for i in range(self.process_list.count()):
            if self.process_list.item(i).checkState() != Qt.Checked:
                all_checked = False
                break

        # 如果全部选中 → 取消
        new_state = Qt.Unchecked if all_checked else Qt.Checked

        for i in range(self.process_list.count()):
            self.process_list.item(i).setCheckState(new_state)

    def add_process_file(self):
        file_dialog = QFileDialog(self)
        file_dialog.setWindowTitle("Select Process Files")
        file_dialog.setFileMode(QFileDialog.ExistingFiles)
        file_dialog.setNameFilter("Excel Files (*.xls *.xlsx)")

        if file_dialog.exec_():

            file_paths = file_dialog.selectedFiles()

            for file_path in file_paths:

                if any(d["process_file"] == file_path for d in self.release_data):
                    continue  # 已存在就跳过

                self.release_data.append({
                    "process_file": file_path,
                    "zc429_files": []
                })

                list_item = QListWidgetItem(file_path)
                list_item.setToolTip(file_path)  # 鼠标悬浮显示完整路径

                # 加 checkbox（和 ENS 一样）
                list_item.setFlags(list_item.flags() | Qt.ItemIsUserCheckable)
                list_item.setCheckState(Qt.Unchecked)

                self.process_list.addItem(list_item)

    def delete_process_file(self):
        rows_to_delete = []

        for row in range(self.process_list.count()):
            item = self.process_list.item(row)

            if item.checkState() == Qt.Checked:
                rows_to_delete.append(row)

        if not rows_to_delete:
            QMessageBox.warning(self, "Warning", "Please select process files to delete.")
            return

        for row in reversed(rows_to_delete):
            self.process_list.takeItem(row)
            del self.release_data[row]

        self.zc429_list.clear()

    def add_429_file(self):
        row = self.process_list.currentRow()

        if row < 0:
            QMessageBox.warning(self, "Warning", "Please select a process file first.")
            return

        file_dialog = QFileDialog(self)
        file_dialog.setWindowTitle("Select ZC429 Files")
        file_dialog.setFileMode(QFileDialog.ExistingFiles)
        file_dialog.setNameFilter("XML Files (*.xml)")

        if file_dialog.exec_():
            self.zc429_list.clear()
            self.release_data[row]["zc429_files"] = []

            for file_path in file_dialog.selectedFiles():

                self.release_data[row]["zc429_files"].append(file_path)

                list_item = QListWidgetItem(file_path)
                list_item.setToolTip(file_path)  # 鼠标悬浮显示完整路径
                self.zc429_list.addItem(list_item)

    def delete_429_file(self):
        row = self.process_list.currentRow()

        if row < 0:
            return

        self.release_data[row]["zc429_files"].clear()

        self.zc429_list.clear()

    def write_release(self):

        if not self.release_data:
            QMessageBox.warning(self, "Warning", "No process file to write.")
            return

        # 检查是否缺少 ZC429
        missing = []
        for data in self.release_data:
            if not data["zc429_files"]:
                missing.append(data["process_file"])

        if missing:
            msg = "The following process files are missing ZC429 files:\n\n"
            msg += "\n".join(missing)
            QMessageBox.warning(self, "Missing XML Files", msg)
            return

        success_list = []
        failed_list = []

        for data in self.release_data:

            process_path = data["process_file"]
            zc429_files = data["zc429_files"]

            try:
                matched, unmatched = self.write_single_process_release(
                    process_path,
                    zc429_files
                )

                success_list.append(
                    f"{process_path} (Matched: {matched}, Unmatched: {unmatched})"
                )

            except Exception as e:
                failed_list.append(f"{process_path} -> {str(e)}")

        msg = f"Release write completed.\n\nSuccess: {len(success_list)} file(s)"

        if success_list:
            msg += "\n\nDetails:\n" + "\n".join(success_list)

        if failed_list:
            msg += f"\n\nFailed: {len(failed_list)} file(s)\n" + "\n".join(failed_list)

        QMessageBox.information(self, "Result", msg)

    def write_single_process_release(self, process_path, zc429_files):

        # === Step1: 解析所有 ZC429 ===
        all_release_info = []

        for file_path in zc429_files:
            try:
                with open(file_path, "rb") as f:
                    all_release_info.extend(parse_zc429_xml(f))
            except Exception as e:
                print(f"Error parsing {file_path}: {e}")

        if not all_release_info:
            raise ValueError("No valid MRN found in ZC429 files.")

        # # 构造 UCR → MRN 映射
        # ucr_to_mrn = dict(all_release_info)

        # 适配后缀是A的情况
        suffix = self.suffix_input.text().strip()

        ucr_to_mrn = {}

        for ucr, mrn in all_release_info:

            original = ucr.strip()
            ucr_to_mrn[original] = mrn  # 先保留原始

            if suffix and original.endswith(suffix):
                cleaned = original[:-len(suffix)]
                if cleaned not in ucr_to_mrn:
                    ucr_to_mrn[cleaned] = mrn

        # === Step2: 打开 Excel（保留格式）===
        wb = load_workbook(process_path)
        ws = wb.active  # 如果固定 sheet 建议改成 wb["Sheet1"]

        tracking_col = 2
        mrn_col = 6

        column_text = self.column_input.text().strip()
        if column_text:

            if not column_text.isdigit():
                QMessageBox.warning(self, "Warning", "Please input a valid column number.")
                return
            mrn_col = int(column_text)

        matched_count = 0
        unmatched_count = 0

        # === Step3: 填充数据（不破坏样式）===
        for row in range(2, ws.max_row + 1):

            tracking_value = str(ws.cell(row=row, column=tracking_col).value).strip()
            if tracking_value in ucr_to_mrn:
                ws.cell(row=row, column=mrn_col).value = ucr_to_mrn[tracking_value]
                matched_count += 1
            else:
                unmatched_count += 1

        wb.save(process_path)

        return matched_count, unmatched_count

    # 还需要ManiFest和flight
    def submit(self):
        # 禁用按钮，改样式、文字
        self.findChild(QPushButton, "confirmbutton").setEnabled(False)
        self.findChild(QPushButton, "confirmbutton").setText("Processing...")

        QApplication.processEvents()

        try:
            # self.json_response = {
            #     "login": "cacesa_test",
            #     "pwd": "password",
            #     "function": "addItemsReleases",
            #     "data": {
            #         "MAWB": "080-47794491",
            #         "flight": "LO1234",
            #         "date": "2022-04-13",
            #         "boxes": [{
            #             "num": "MIX26831616GH",
            #             "items": [
            #                 {"parcel": "2137935414089", "release": "22PL443020AA111111"},
            #                 {"parcel": "1780348473216", "release": "22PL443020AA222222"},
            #                 {"parcel": "3202903066746", "release": "22PL443020AA333333"},
            #                 {"parcel": "8828190634800", "release": "22PL443020AA444444"},
            #                 {"parcel": "580503663265", "release": "22PL443020AA444444"},
            #                 {"parcel": "7361849575125", "release": "22PL443020AA444444"}
            #             ]
            #         }, {
            #             "num": "MIX659718510GH",
            #             "items": [
            #                 {"parcel": "4601252975875", "release": "22PL443020A555555"},
            #                 {"parcel": "71306329993", "release": "22PL443020A555555"},
            #                 {"parcel": "1768882841110", "release": "22PL443020AA666666"},
            #                 {"parcel": "13193951760210", "release": "22PL443020AA666666"},
            #                 {"parcel": "1914756791100", "release": "22PL443020AA444444"},
            #                 {"parcel": "12532753191420", "release": "22PL443020AA444444"}
            #             ]
            #         }]
            #     }
            # }

            request_response = self.api.add_items_releases()
            try:
                if request_response.status_code == 200:
                    response_json = request_response.json()
                    print("get response successfully.")
                    if response_json.get('error'):
                        err_msg = response_json.get('error', 'Unknown error')
                        print(f"Add items releases failed: {err_msg}")
                        QMessageBox.critical(self, "Error", f"Add items releases failed! {err_msg}")
                    else:
                        print("Add items releases successfully.")
                        QMessageBox.information(self, "Success", "Add items releases successfully!")
                        # self.finished = True
                        self.save_button.setEnabled(True)
                        # self.finished_successful.emit()
                else:
                    print(f"Failed to Add items releases. Status code: {request_response.status_code}")
                    print("Raw response text:", request_response.text)
                    QMessageBox.critical(self, "Error",
                                         f"Failed to Add items releases. Status code: {request_response.status_code}")

            except AttributeError:
                print("Error: Response object does not have a status_code attribute.")
                QMessageBox.critical(self, "Error", "Invalid response object received.")

        finally:
            # 恢复按钮状态
            self.findChild(QPushButton, "confirmbutton").setEnabled(True)
            self.findChild(QPushButton, "confirmbutton").setText("Submit")
            QApplication.processEvents()

    def on_cancel(self):
        # 创建提示框，确认取消操作
        reply = QMessageBox.question(self, 'Confirm Cancel',
                                     "This action will discard all changes and return to the first page. Are you sure you want to continue?",
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

        if reply == QMessageBox.Yes:
            self.cancel_requested.emit()  # 发送信号，通知父窗口取消操作
        else:
            print("User canceled the action.")

    def save_release(self):
        self.save_box_parcel_release_excel(self.json_response)

    def save_box_parcel_release_excel(parent_widget, json_response: dict):
        try:
            boxes = json_response.get("data", {}).get("boxes", [])
            if not boxes:
                QMessageBox.warning(parent_widget, "No Data", "No 'boxes' found in response.")
                return

            # 构造数据行
            data = []
            for box in boxes:
                boxnumber = box.get("num", "")
                items = box.get("items", [])
                for idx, item in enumerate(items):
                    parcel = item.get("parcel", "")
                    release = item.get("release", "")
                    data.append([
                        boxnumber if idx == 0 else "",  # 后续行 boxnumber 为空
                        parcel,
                        release
                    ])

            df = pd.DataFrame(data, columns=["boxnumber", "parcel", "release"])

            default_filename = f"{global_state.flight}_{global_state.carrier}_releases.xlsx"

            # 用户选择保存路径
            save_path, _ = QFileDialog.getSaveFileName(
                parent_widget,
                "Save File As",
                default_filename,
                "Excel Files (*.xlsx)"
            )

            if not save_path:
                return

            if not save_path.endswith(".xlsx"):
                save_path += ".xlsx"

            # 保存 Excel
            df.to_excel(save_path, index=False)

            QMessageBox.information(parent_widget, "Success", f"File successfully saved to:\n{save_path}")

        except Exception as e:
            QMessageBox.critical(parent_widget, "Error", f"An error occurred:\n{e}")

    def get_release_mrn(self):
        dialog = ReleaseModeDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            mode = dialog.selected_mode
            if mode == '>150':
                MultiFileUploadDialog(self).exec_()
            else:
                OldmultiFileUploadDialog(self).exec_()
