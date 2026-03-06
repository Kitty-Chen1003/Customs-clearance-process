import openpyxl
from PyQt5.QtGui import QFont
from openpyxl.styles import Alignment
from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QLabel, QLineEdit, QPushButton,
    QHBoxLayout, QFileDialog, QMessageBox, QScrollArea,
    QListWidget, QListWidgetItem, QTableWidget, QTableWidgetItem,
    QAbstractItemView, QHeaderView, QFrame, QApplication, QSizePolicy
)


class AddArn(QWidget):
    cancel_requested = pyqtSignal()

    def __init__(self, api):
        super().__init__()
        self.column_input = None
        self.step_config = None
        self.api = api

        # ===== 核心数据 =====
        self.arn_data = []  # [{ arn: str, process_files: [] }]

        self.outer_layout = None
        self.scroll_area = None
        self.inner_widget = None
        self.inner_layout = None
        self.scroll_area_label = None
        self.label_status = None

        self.arn_input = None
        self.arn_list = None
        self.process_file_list = None

        self.init_ui()

    def set_step_config(self, step_config: dict):
        """设置字段名和提示"""
        self.step_config = step_config
        self.update_hint_label()

    def update_hint_label(self):
        """在页面标题下方增加提示"""
        if not self.step_config or not hasattr(self, "inner_layout"):
            return

        # 删除已有提示标签（如果存在）
        for i in reversed(range(self.inner_layout.count())):
            widget = self.inner_layout.itemAt(i).widget()
            if isinstance(widget, QLabel) and getattr(widget, "is_hint_label", False):
                self.inner_layout.removeWidget(widget)
                widget.deleteLater()

        hint_text = self.step_config.get("hint", "")
        if hint_text:
            hint_label = QLabel(hint_text)
            hint_label.setObjectName("hint")
            hint_label.setAlignment(Qt.AlignCenter)
            hint_label.is_hint_label = True  # 标记为提示
            # 找到标题位置后插入到其下方
            # 假设标题总是 inner_layout 的第一个 QLabel
            for i in range(self.inner_layout.count()):
                widget = self.inner_layout.itemAt(i).widget()
                if isinstance(widget, QLabel) and widget.objectName() == "title":
                    self.inner_layout.insertWidget(i + 1, hint_label)
                    break

    def init_ui(self):

        self.build_outer_layout()

        self.build_scroll_area()

        self.build_inner_layout()

        self.build_title()

        self.build_arn_input()

        self.build_column_input()

        self.build_arn_process_panel()

        self.build_action_buttons()

    def build_outer_layout(self):

        # 外层布局（用于居中）
        self.outer_layout = QVBoxLayout(self)
        self.outer_layout.setContentsMargins(0, 0, 0, 0)
        self.outer_layout.setAlignment(Qt.AlignCenter)

    def build_scroll_area(self):
        # 创建 QScrollArea，并设置属性
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)  # 让 scroll area 根据内容自适应
        # self.scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)  # 横向不要滚动条（可选）
        self.scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)  # 纵向滚动条按需出现

        self.outer_layout.addWidget(self.scroll_area)

    def build_inner_layout(self):
        # 内层容器和布局（承载内容）
        self.inner_widget = QWidget()
        self.inner_layout = QVBoxLayout(self.inner_widget)
        self.inner_layout.setContentsMargins(100, 50, 100, 50)
        self.inner_layout.setSpacing(20)
        self.inner_layout.setAlignment(Qt.AlignTop | Qt.AlignHCenter)  # 控制内容居中靠上

        self.scroll_area.setWidget(self.inner_widget)

    def build_title(self):
        # 标题
        title = QLabel("AddARN")
        title.setObjectName("title")
        title.setAlignment(Qt.AlignCenter)
        self.inner_layout.addWidget(title)

    def build_arn_input(self):
        # 输入框
        arn_layout = QVBoxLayout()

        arn_label = QLabel("ARN Number:")
        arn_label.setObjectName("normal")
        self.arn_input = QLineEdit()
        # self.flight_input.setMaximumWidth(800)
        # self.flight_input.setMinimumWidth(300)
        self.arn_input.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        add_arn_button = QPushButton("Add ARN")
        add_arn_button.setFont(QFont("Microsoft YaHei", 12))
        # add_fc_button.setMaximumWidth(800)
        add_arn_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        add_arn_button.setStyleSheet("padding: 10px;")
        add_arn_button.clicked.connect(self.add_arn_number)

        arn_layout.addWidget(arn_label)
        arn_layout.addWidget(self.arn_input)
        arn_layout.addSpacing(10)
        arn_layout.addWidget(add_arn_button)
        arn_layout.setAlignment(Qt.AlignHCenter)  # 整个控件居中

        self.inner_layout.addLayout(arn_layout)
        self.inner_layout.addSpacing(10)

    def build_column_input(self):
        column_layout = QVBoxLayout()

        column_label = QLabel("Column Number:")
        column_label.setObjectName("normal")
        self.column_input = QLineEdit()
        self.column_input.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        column_layout.addWidget(column_label)
        column_layout.addWidget(self.column_input)
        column_layout.setAlignment(Qt.AlignHCenter)  # 整个控件居中

        self.inner_layout.addLayout(column_layout)
        self.inner_layout.addSpacing(10)

    def build_arn_process_panel(self):

        arn_process_layout = QHBoxLayout()

        # 左侧 arn 列表
        left_layout = QVBoxLayout()
        arn_label = QLabel("ARN:")
        arn_label.setObjectName("normal")

        self.arn_list = QTableWidget()
        self.arn_list.setColumnCount(1)
        self.arn_list.setHorizontalHeaderLabels(["ARN"])

        self.arn_list.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.arn_list.setMinimumHeight(300)

        self.arn_list.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.arn_list.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.arn_list.horizontalHeader().setStretchLastSection(True)
        self.arn_list.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.arn_list.verticalHeader().setVisible(False)
        self.arn_list.setSelectionMode(QAbstractItemView.SingleSelection)  # 单选模式

        self.arn_list.setStyleSheet("""
            QTableWidget {
                background-color: #ffffff;
                gridline-color: #ccc;
                border: 1px solid #aaa;
            }
            QHeaderView::section {
                background-color: #0078d7;
                color: white;
                padding: 6px 12px;
                border: 1px solid #aaa;
            }
            QTableWidget::item:selected {
                background-color: #87cefa;
                color: black;
            }
        """)
        self.arn_list.itemClicked.connect(self.show_processes)

        # TODO:按钮设置样式
        arn_btn_layout = QHBoxLayout()

        table_select_all_btn = QPushButton("Select All / Unselect All")
        table_select_all_btn.setObjectName("fcAddButton")
        table_select_all_btn.clicked.connect(self.table_toggle_selects)

        del_arn_btn = QPushButton("Delete Selected ARN")
        del_arn_btn.setObjectName("fcDelButton")
        del_arn_btn.clicked.connect(self.delete_selected_arn)

        arn_btn_layout.addWidget(table_select_all_btn)
        arn_btn_layout.addWidget(del_arn_btn)

        left_layout.addWidget(arn_label)
        left_layout.addWidget(self.arn_list)
        left_layout.addLayout(arn_btn_layout)

        # 右侧 manifest 文件列表
        right_layout = QVBoxLayout()
        process_label = QLabel("process file:")
        process_label.setObjectName("normal")
        self.process_file_list = QListWidget()
        self.process_file_list.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.process_file_list.setMinimumHeight(300)

        # 支持多选
        self.process_file_list.setSelectionMode(QAbstractItemView.NoSelection)

        self.process_file_list.setStyleSheet("""
            QListWidget {
                background-color: #f9f9f9;
                border: 1px solid #ccc;
                padding: 5px;
            }
            QListWidget::item {
                padding: 5px;
            }
            QListWidget::item:selected {
                background-color: #87cefa;
                color: black;
            }
        """)

        # manifest 操作按钮
        process_btn_layout = QHBoxLayout()

        select_all_btn = QPushButton("Select All / Unselect All")
        select_all_btn.setObjectName("fcAddButton")
        select_all_btn.clicked.connect(self.toggle_selects)

        add_process_btn = QPushButton("Add Process File")
        add_process_btn.clicked.connect(self.add_process)

        del_process_btn = QPushButton("Delete Selected Process File")
        del_process_btn.clicked.connect(self.delete_selected_process)

        add_process_btn.setObjectName("fcAddButton")
        del_process_btn.setObjectName("fcDelButton")

        process_btn_layout.addWidget(select_all_btn)
        process_btn_layout.addWidget(add_process_btn)
        process_btn_layout.addWidget(del_process_btn)

        right_layout.addWidget(process_label)
        right_layout.addWidget(self.process_file_list)
        right_layout.addLayout(process_btn_layout)

        arn_process_group = QFrame()
        arn_process_group.setObjectName("fcFrame")
        arn_process_group.setFrameShape(QFrame.NoFrame)
        arn_process_group.setLayout(arn_process_layout)

        # fc_manifest_layout.addStretch()
        arn_process_layout.addLayout(left_layout)
        arn_process_layout.addSpacing(50)
        arn_process_layout.addLayout(right_layout)
        # fc_manifest_layout.addStretch()

        arn_process_layout.setStretchFactor(left_layout, 1)
        arn_process_layout.setStretchFactor(right_layout, 2)

        # self.inner_layout.addLayout(fc_manifest_layout)
        self.inner_layout.addWidget(arn_process_group)

        self.inner_layout.addSpacing(20)

    def build_action_buttons(self):
        write_arn_button = QPushButton("Write ARN to process file")
        write_arn_button.setFont(QFont("Microsoft YaHei", 12))
        write_arn_button.setStyleSheet("padding: 10px;")
        write_arn_button.clicked.connect(self.write_arn)

        self.inner_layout.addWidget(write_arn_button)
        self.inner_layout.addSpacing(20)

        # 按钮水平布局（发送按钮和取消按钮并排）
        button_layout = QHBoxLayout()

        # 发送按钮
        submit_button = QPushButton("Submit")
        submit_button.setObjectName("confirmbutton")
        submit_button.clicked.connect(self.submit)

        # 取消按钮
        cancel_button = QPushButton("Cancel")
        cancel_button.setObjectName("cancelButton")
        cancel_button.clicked.connect(self.on_cancel)  # 连接到提示框槽函数

        button_layout.addWidget(submit_button)
        button_layout.addWidget(cancel_button)

        # 将按钮布局添加到主布局中
        self.inner_layout.addLayout(button_layout)
        self.inner_layout.addSpacing(20)

        # clear_button = QPushButton("Clear")
        # clear_button.setFont(QFont("Microsoft YaHei", 12))
        # clear_button.setStyleSheet("padding: 10px;")
        # clear_button.clicked.connect(self.reset)
        # self.inner_layout.addWidget(clear_button)

        # 底部按钮布局
        self.inner_layout.addStretch()

    # ================= Data Ops =================

    def add_arn_number(self):
        arn = self.arn_input.text().strip()
        if not arn:
            QMessageBox.warning(self, "Warning", "ARN number must be filled")
            return

        if any(a["arn"] == arn for a in self.arn_data):
            QMessageBox.warning(self, "Warning", "ARN already exists.")
            return

        self.arn_data.append({'arn': arn, 'process_files': []})

        row = self.arn_list.rowCount()
        self.arn_list.insertRow(row)

        arn_item = QTableWidgetItem(arn)
        arn_item.setFlags(arn_item.flags() | Qt.ItemIsUserCheckable)
        arn_item.setCheckState(Qt.Unchecked)

        self.arn_list.setItem(row, 0, arn_item)

        self.arn_input.clear()

    def table_toggle_selects(self):
        if self.arn_list.rowCount() == 0:
            return

        # 判断是否全部勾选
        all_checked = True
        for row in range(self.arn_list.rowCount()):
            item = self.arn_list.item(row, 0)  # 第一列是 checkbox
            if item.checkState() != Qt.Checked:
                all_checked = False
                break

        new_state = Qt.Unchecked if all_checked else Qt.Checked

        for row in range(self.arn_list.rowCount()):
            self.arn_list.item(row, 0).setCheckState(new_state)

    def toggle_selects(self):
        if self.process_file_list.count() == 0:
            return

        # 判断是否已经全部选中
        all_checked = True
        for i in range(self.process_file_list.count()):
            if self.process_file_list.item(i).checkState() != Qt.Checked:
                all_checked = False
                break

        # 如果全部选中 → 取消
        new_state = Qt.Unchecked if all_checked else Qt.Checked

        for i in range(self.process_file_list.count()):
            self.process_file_list.item(i).setCheckState(new_state)

    def delete_selected_arn(self):
        rows_to_delete = []

        for row in range(self.arn_list.rowCount()):
            item = self.arn_list.item(row, 0)

            if item.checkState() == Qt.Checked:
                rows_to_delete.append(row)

        if not rows_to_delete:
            QMessageBox.warning(self, "Warning", "Please select arn to delete")
            return

        for row in reversed(rows_to_delete):
            self.arn_list.removeRow(row)
            self.arn_data.pop(row)

        self.process_file_list.clear()  # 同时清空右侧 manifest

    def show_processes(self, item):
        self.process_file_list.clear()

        row = self.arn_list.row(item)  # ✅ QTableWidgetItem → int
        if row < 0 or row >= len(self.arn_data):
            return

        for process_path in self.arn_data[row]["process_files"]:
            list_item = QListWidgetItem(process_path)
            list_item.setToolTip(process_path)
            list_item.setFlags(list_item.flags() | Qt.ItemIsUserCheckable)
            list_item.setCheckState(Qt.Unchecked)
            self.process_file_list.addItem(list_item)

    def add_process(self):
        selected_arn_items = self.arn_list.selectedItems()
        if not selected_arn_items:
            QMessageBox.warning(self, "Warning", "Please select an ARN first")
            return

        file_dialog = QFileDialog(self)
        file_dialog.setWindowTitle('Select Process Files')
        file_dialog.setFileMode(QFileDialog.ExistingFiles)  # 只允许选择一个文件
        file_dialog.setNameFilter("Excel Files (*.xls *.xlsx)")

        if file_dialog.exec_():
            arn_idx = self.arn_list.currentRow()
            for file_path in file_dialog.selectedFiles():
                if file_path in self.arn_data[arn_idx]['process_files']:
                    continue
                self.arn_data[arn_idx]['process_files'].append(file_path)

                list_item = QListWidgetItem(file_path)
                list_item.setToolTip(file_path)  # 鼠标悬浮显示完整路径
                list_item.setFlags(list_item.flags() | Qt.ItemIsUserCheckable)  # 可勾选
                list_item.setCheckState(Qt.Unchecked)
                self.process_file_list.addItem(list_item)

    def delete_selected_process(self):
        arn_idx = self.arn_list.currentRow()
        if arn_idx < 0:
            return
        # 倒序删除，避免索引错乱
        for row in reversed(range(self.process_file_list.count())):
            item = self.process_file_list.item(row)
            if item.checkState() == Qt.Checked:
                self.process_file_list.takeItem(row)
                self.arn_data[arn_idx]['process_files'].pop(row)

    # ================= Write Logic =================

    def write_arn(self):
        arn_col = "MRN(ARN) from ICS2\n EU based on AWB - F21"

        if not self.arn_data:
            QMessageBox.warning(self, "Warning", "No ARN data to write.")
            return

        wrote_any = False  # 是否真的写入过

        for entry in self.arn_data:
            arn = entry.get("arn")
            process_files = entry.get("process_files", [])

            if not arn or not process_files:
                continue  # 跳过没有文件的 ARN

            for path in process_files:
                try:
                    wb = openpyxl.load_workbook(path)
                    ws = wb.active

                    headers = [c.value for c in ws[1]]
                    if arn_col not in headers:
                        QMessageBox.warning(
                            self,
                            "Missing Column",
                            f"{arn_col}\nnot found in:\n{path}"
                        )
                        continue

                    col = headers.index(arn_col) + 1

                    #TODO 注意
                    column_text = self.column_input.text().strip()
                    if column_text:

                        if not column_text.isdigit():
                            QMessageBox.warning(self, "Warning", "Please input a valid column number.")
                            return
                        col = int(column_text)

                    for r in range(2, ws.max_row + 1):
                        cell = ws.cell(r, col)
                        cell.value = arn

                    wb.save(path)
                    wrote_any = True

                except Exception as e:
                    QMessageBox.critical(self, "Error", f"Failed to write MRN(ARN):\n{e}")
                    return

        if wrote_any:
            QMessageBox.information(self, "Success", "MRN(ARN) written successfully!")
        else:
            QMessageBox.warning(
                self,
                "Warning",
                "No ARN was written. Please check ARN and process file selection."
            )

    def submit(self):
        self.findChild(QPushButton, "confirmbutton").setEnabled(False)
        self.findChild(QPushButton, "confirmbutton").setText("Processing...")

        QApplication.processEvents()

        try:

            request_response = self.api.add_arn()
            # self.finished = True
            # self.finished_successful.emit()
            try:
                if request_response.status_code == 200:
                    response_json = request_response.json()
                    print("get response successfully.")
                    if response_json.get('success', False):
                        print("Add ARN successfully.")
                        QMessageBox.information(self, "Success", "Add ARN successfully!")
                        # self.finished = True
                        # self.finished_successful.emit()  # 发送信号，通知父窗口更新数据
                    else:
                        err_msg = response_json.get('error', 'Unknown error')
                        print(f"Add ARN failed: {err_msg}")
                        QMessageBox.critical(self, "Error", f"Add ARN failed! {err_msg}")
                else:
                    print(f"Failed to add ARN. Status code: {request_response.status_code}")
                    print("Raw response text:", request_response.text)
                    QMessageBox.critical(self, "Error",
                                         f"Failed to add ARN. Status code: {request_response.status_code}")
            except AttributeError:
                print("Error: Response object does not have a status_code attribute.")
                QMessageBox.critical(self, "Error", "Invalid response object received.")

        finally:
            # 恢复按钮状态
            self.findChild(QPushButton, "confirmbutton").setEnabled(True)
            self.findChild(QPushButton, "confirmbutton").setText("Submit")
            QApplication.processEvents()

    def on_cancel(self):
        # 创建提示框，确认取消操作
        reply = QMessageBox.question(self, 'Confirm Cancel',
                                     "This action will discard all changes and return to the first page. Are you sure you want to continue?",
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

        if reply == QMessageBox.Yes:
            self.cancel_requested.emit()  # 发送信号，通知父窗口取消操作
        else:
            print("User canceled the action.")
