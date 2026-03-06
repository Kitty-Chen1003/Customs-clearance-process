from copy import copy

import pandas as pd
from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtGui import QFont
from PyQt5.QtWidgets import QWidget, QVBoxLayout, QLabel, QLineEdit, QSizePolicy, QPushButton, QHBoxLayout, QFileDialog, \
    QMessageBox, QScrollArea, QApplication, QListWidget, QAbstractItemView, QFrame, QListWidgetItem
from openpyxl.reader.excel import load_workbook

import global_state
from http_client import get_dsk_request


class GetDsk(QWidget):
    # finished_successful = pyqtSignal()
    cancel_requested = pyqtSignal()

    def __init__(self, api):

        super().__init__()

        self.step_config = None
        self.api = api
        # self.finished = False

        self.dsk_data = []  # [{ manifest_file: str, dsk_file: str }]

        self.outer_layout = None
        self.scroll_area = None
        self.inner_widget = None
        self.inner_layout = None
        self.scroll_area_label = None
        self.label_status = None

        self.manifest_list = None
        self.dsk_list = None

        self.label_mapping = None
        self.scroll_area_mapping = None

        self.response_json = None
        self.save_button = None

        self.init_ui()
        
    def set_step_config(self, step_config: dict):
        """设置字段名和提示"""
        self.step_config = step_config
        self.update_hint_label()

    def update_hint_label(self):
        """在页面标题下方增加提示"""
        if not self.step_config or not hasattr(self, "inner_layout"):
            return

        # 删除已有提示标签（如果存在）
        for i in reversed(range(self.inner_layout.count())):
            widget = self.inner_layout.itemAt(i).widget()
            if isinstance(widget, QLabel) and getattr(widget, "is_hint_label", False):
                self.inner_layout.removeWidget(widget)
                widget.deleteLater()

        hint_text = self.step_config.get("hint", "")
        if hint_text:
            hint_label = QLabel(hint_text)
            hint_label.setObjectName("hint")
            hint_label.setAlignment(Qt.AlignCenter)
            hint_label.is_hint_label = True  # 标记为提示
            # 找到标题位置后插入到其下方
            # 假设标题总是 inner_layout 的第一个 QLabel
            for i in range(self.inner_layout.count()):
                widget = self.inner_layout.itemAt(i).widget()
                if isinstance(widget, QLabel) and widget.objectName() == "title":
                    self.inner_layout.insertWidget(i + 1, hint_label)
                    break

    def init_ui(self):

        self.build_outer_layout()

        self.build_scroll_area()

        self.build_inner_layout()

        self.build_title()

        self.build_process_ens_panel()

        self.build_action_buttons()

    def build_outer_layout(self):

        # 外层布局（用于居中）
        self.outer_layout = QVBoxLayout(self)
        self.outer_layout.setContentsMargins(0, 0, 0, 0)
        self.outer_layout.setAlignment(Qt.AlignCenter)

    def build_scroll_area(self):
        # 创建 QScrollArea，并设置属性
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)  # 让 scroll area 根据内容自适应
        # self.scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)  # 横向不要滚动条（可选）
        self.scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)  # 纵向滚动条按需出现

        self.outer_layout.addWidget(self.scroll_area)

    def build_inner_layout(self):
        # 内层容器和布局（承载内容）
        self.inner_widget = QWidget()
        self.inner_layout = QVBoxLayout(self.inner_widget)
        self.inner_layout.setContentsMargins(100, 50, 100, 50)
        self.inner_layout.setSpacing(20)
        self.inner_layout.setAlignment(Qt.AlignTop | Qt.AlignHCenter)  # 控制内容居中靠上

        self.scroll_area.setWidget(self.inner_widget)

    def build_title(self):
        # 标题
        title = QLabel("GetDSK")
        title.setObjectName("title")
        title.setAlignment(Qt.AlignCenter)
        self.inner_layout.addWidget(title)
        # self.inner_layout.addSpacing(50)

    def build_process_ens_panel(self):

        manifest_dsk_layout = QHBoxLayout()

        # 左侧 arn 列表
        left_layout = QVBoxLayout()
        manifest_label = QLabel("Manifest Files:")
        manifest_label.setObjectName("normal")

        self.manifest_list = QListWidget()
        self.manifest_list.setSelectionMode(QAbstractItemView.SingleSelection)

        self.manifest_list.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.manifest_list.setMinimumHeight(300)

        self.manifest_list.setStyleSheet("""
            QListWidget {
                background-color: #f9f9f9;
                border: 1px solid #ccc;
                padding: 5px;
            }
            QListWidget::item {
                padding: 5px;
            }
            QListWidget::item:selected {
                background-color: #87cefa;
                color: black;
            }
        """)

        self.manifest_list.itemClicked.connect(self.show_dsk)

        # TODO:按钮设置样式
        manifest_btn_layout = QHBoxLayout()

        select_all_btn = QPushButton("Select All / Unselect All")
        select_all_btn.setObjectName("fcAddButton")
        select_all_btn.clicked.connect(self.toggle_selects)

        add_manifest_btn = QPushButton("Add Manifest File")
        add_manifest_btn.setObjectName("fcAddButton")
        add_manifest_btn.clicked.connect(self.add_manifest_file)

        del_manifest_btn = QPushButton("Delete Manifest File")
        del_manifest_btn.setObjectName("fcDelButton")
        del_manifest_btn.clicked.connect(self.delete_manifest_file)

        manifest_btn_layout.addWidget(select_all_btn)
        manifest_btn_layout.addWidget(add_manifest_btn)
        manifest_btn_layout.addWidget(del_manifest_btn)

        left_layout.addWidget(manifest_label)
        left_layout.addWidget(self.manifest_list)
        left_layout.addLayout(manifest_btn_layout)

        # 右侧 mapping 文件列表
        right_layout = QVBoxLayout()
        dsk_label = QLabel("Dsk File:")
        dsk_label.setObjectName("normal")

        self.dsk_list = QListWidget()
        self.dsk_list.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.dsk_list.setMinimumHeight(300)
        self.dsk_list.setSelectionMode(QAbstractItemView.NoSelection)  # 单选模式

        self.dsk_list.setStyleSheet("""
            QListWidget {
                background-color: #f9f9f9;
                border: 1px solid #ccc;
                padding: 5px;
            }
            QListWidget::item {
                padding: 5px;
            }
            QListWidget::item:selected {
                background-color: #87cefa;
                color: black;
            }
        """)

        # mapping 操作按钮
        dsk_btn_layout = QHBoxLayout()
        add_dsk_btn = QPushButton("Add Dsk File")
        add_dsk_btn.clicked.connect(self.add_dsk_file)

        del_dsk_btn = QPushButton("Delete Dsk File")
        del_dsk_btn.clicked.connect(self.delete_dsk_file)

        add_dsk_btn.setObjectName("fcAddButton")
        del_dsk_btn.setObjectName("fcDelButton")

        dsk_btn_layout.addWidget(add_dsk_btn)
        dsk_btn_layout.addWidget(del_dsk_btn)

        right_layout.addWidget(dsk_label)
        right_layout.addWidget(self.dsk_list)
        right_layout.addLayout(dsk_btn_layout)

        manifest_dsk_group = QFrame()
        manifest_dsk_group.setObjectName("fcFrame")
        manifest_dsk_group.setFrameShape(QFrame.NoFrame)
        manifest_dsk_group.setLayout(manifest_dsk_layout)

        # fc_manifest_layout.addStretch()
        manifest_dsk_layout.addLayout(left_layout)
        manifest_dsk_layout.addSpacing(50)
        manifest_dsk_layout.addLayout(right_layout)
        # fc_manifest_layout.addStretch()

        manifest_dsk_layout.setStretchFactor(left_layout, 1)
        manifest_dsk_layout.setStretchFactor(right_layout, 1)

        # self.inner_layout.addLayout(fc_manifest_layout)
        self.inner_layout.addWidget(manifest_dsk_group)

        self.inner_layout.addSpacing(20)

    def reset(self):
        pass

    def build_action_buttons(self):
        write_btn = QPushButton("Write Dsk to Manifest File")
        write_btn.setFont(QFont("Microsoft YaHei", 12))
        write_btn.setStyleSheet("padding: 10px;")
        write_btn.clicked.connect(self.write_dsk)

        self.inner_layout.addWidget(write_btn)
        self.inner_layout.addSpacing(20)

        btn_layout = QHBoxLayout()

        submit_btn = QPushButton("Submit")
        submit_btn.setObjectName("confirmbutton")
        submit_btn.clicked.connect(self.submit)

        cancel_btn = QPushButton("Cancel")
        cancel_btn.setObjectName("cancelButton")
        cancel_btn.clicked.connect(self.on_cancel)

        btn_layout.addWidget(submit_btn)
        btn_layout.addWidget(cancel_btn)

        self.inner_layout.addLayout(btn_layout)
        self.inner_layout.addSpacing(20)

        self.save_button = QPushButton("Save")
        self.save_button.setFont(QFont("Microsoft YaHei", 12))
        self.save_button.setStyleSheet("padding: 10px;")
        self.save_button.clicked.connect(self.save)
        self.save_button.setEnabled(False)

        self.inner_layout.addWidget(self.save_button)
        self.inner_layout.addStretch()

    def show_dsk(self, item):
        self.dsk_list.clear()

        row = self.manifest_list.row(item)  # ✅ QTableWidgetItem → int
        if row < 0 or row >= len(self.dsk_data):
            return

        dsk_file = self.dsk_data[row].get("dsk_file")

        if dsk_file:
            list_item = QListWidgetItem(dsk_file)
            list_item.setToolTip(dsk_file)
            self.dsk_list.addItem(list_item)

    def toggle_selects(self):
        if self.manifest_list.count() == 0:
            return

        # 判断是否已经全部选中
        all_checked = True
        for i in range(self.manifest_list.count()):
            if self.manifest_list.item(i).checkState() != Qt.Checked:
                all_checked = False
                break

        # 如果全部选中 → 取消
        new_state = Qt.Unchecked if all_checked else Qt.Checked

        for i in range(self.manifest_list.count()):
            self.manifest_list.item(i).setCheckState(new_state)

    def add_manifest_file(self):
        file_dialog = QFileDialog(self)
        file_dialog.setWindowTitle("Select Manifest Files")
        file_dialog.setFileMode(QFileDialog.ExistingFiles)
        file_dialog.setNameFilter("Excel Files (*.xls *.xlsx)")

        if file_dialog.exec_():

            file_paths = file_dialog.selectedFiles()

            for file_path in file_paths:
                if any(d["manifest_file"] == file_path for d in self.dsk_data):
                    continue  # 已存在就跳过

                self.dsk_data.append({
                    "manifest_file": file_path,
                    "dsk_file": ""
                })

                list_item = QListWidgetItem(file_path)
                list_item.setToolTip(file_path)  # 鼠标悬浮显示完整路径

                # 添加 checkbox
                list_item.setFlags(list_item.flags() | Qt.ItemIsUserCheckable)
                list_item.setCheckState(Qt.Unchecked)

                self.manifest_list.addItem(list_item)

    def delete_manifest_file(self):
        rows_to_delete = []

        for row in range(self.manifest_list.count()):
            item = self.manifest_list.item(row)

            if item.checkState() == Qt.Checked:
                rows_to_delete.append(row)

        if not rows_to_delete:
            QMessageBox.warning(self, "Warning", "Please select manifest files to delete.")
            return

        for row in reversed(rows_to_delete):
            self.manifest_list.takeItem(row)
            del self.dsk_data[row]

        self.dsk_list.clear()


    def add_dsk_file(self):
        row = self.manifest_list.currentRow()

        if row < 0:
            QMessageBox.warning(self, "Warning", "Please select a manifest file first.")
            return

        file_dialog = QFileDialog(self)
        file_dialog.setWindowTitle("Select Dsk File")
        file_dialog.setFileMode(QFileDialog.ExistingFile)
        file_dialog.setNameFilter("Excel Files (*.xls *.xlsx)")

        if file_dialog.exec_():
            file_path = file_dialog.selectedFiles()[0]

            self.dsk_data[row]["dsk_file"] = file_path

            self.dsk_list.clear()
            list_item = QListWidgetItem(file_path)
            list_item.setToolTip(file_path)  # 鼠标悬浮显示完整路径
            self.dsk_list.addItem(list_item)

    def delete_dsk_file(self):
        row = self.manifest_list.currentRow()

        if row < 0:
            return

        self.dsk_data[row]["dsk_file"] = ""

        self.dsk_list.clear()

    def write_dsk(self):
        if not self.dsk_data:
            QMessageBox.warning(self, "Warning", "No manifest file to write.")
            return

            # 检查是否有缺少 mapping 的
        missing = [data["manifest_file"] for data in self.dsk_data if not data.get("dsk_file")]

        if missing:
            msg = "The following manifest files have no dsk file:\n\n"
            msg += "\n".join(missing)

            QMessageBox.warning(self, "Missing dsk File", msg)
            return

        # 全部都有 mapping → 开始写 ENS
        success_list = []
        failed_list = []

        for data in self.dsk_data:

            manifest_path = data["manifest_file"]
            dsk_path = data["dsk_file"]
            try:

                self.write_manifest_dsk(manifest_path, dsk_path)
                success_list.append(manifest_path)

            except Exception as e:
                failed_list.append(f"{manifest_path} -> {str(e)}")

        # 构造提示信息
        msg = f"Dsk write completed.\n\nSuccess: {len(success_list)} file(s)"
        if failed_list:
            msg += f"\nFailed: {len(failed_list)} file(s)\n" + "\n".join(failed_list)

        QMessageBox.information(self, "Result", msg)

    def write_manifest_dsk(self, manifest_path, dsk_path):

        # === 读取 mapping 文件 ===

        # mapping_dict = {
        #     (row['Container code'], row['Tracking code']): row['F26 MRN']
        #     for _, row in mapping_df.iterrows()
        # }
        try:
            mapping_df = pd.read_excel(dsk_path, dtype=str).fillna('')
            mapping_dict = dict(
                zip(
                    zip(mapping_df['Hawb Number'], mapping_df['Item No.']),
                    mapping_df['Dsk']
                )
            )
        except KeyError as e:
            missing_col = e.args[0]  # 会是 'Container code' 或 'Tracking code' 或 'F26 MRN'
            raise ValueError(f"Dsk file {dsk_path} is missing required column: {missing_col}") from e

        # === 打开 process 文件 ===
        wb = load_workbook(manifest_path)
        ws = wb.active

        header = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val:
                header[val.strip()] = col

        box_candidates = [
            "Box Number",
            """Box Number/
Numer Carton"""
        ]

        box_idx = None
        for name in box_candidates:
            if name in header:
                box_idx = header[name]
                break

        tracking_candidates = [
            "TrackingNumber",
            """Parcel Number/
Numer paczki"""
        ]

        tracking_idx = None
        for name in tracking_candidates:
            if name in header:
                tracking_idx = header[name]
                break

        if box_idx is None or tracking_idx is None:
            raise ValueError("Cannot find Box or Tracking column in manifest file")

        if "Dsk" in header:
            f26_idx = header["Dsk"]
        else:
            # 没有就新增
            last_col = max(header.values())
            ref_cell = ws.cell(row=1, column=last_col)

            f26_idx = last_col + 1
            new_cell = ws.cell(row=1, column=f26_idx, value="Dsk")

            new_cell.font = copy(ref_cell.font)
            new_cell.border = copy(ref_cell.border)
            new_cell.fill = copy(ref_cell.fill)
            new_cell.number_format = copy(ref_cell.number_format)
            new_cell.protection = copy(ref_cell.protection)
            new_cell.alignment = copy(ref_cell.alignment)

        # 填充
        for row in range(2, ws.max_row + 1):
            box_val = ws.cell(row=row, column=box_idx).value
            tracking_val = ws.cell(row=row, column=tracking_idx).value
            key = (str(box_val), str(tracking_val))
            ws.cell(row=row, column=f26_idx, value=mapping_dict.get(key, ''))

        # 保存结果
        wb.save(manifest_path)

    def submit(self):
        self.findChild(QPushButton, "confirmbutton").setEnabled(False)
        self.findChild(QPushButton, "confirmbutton").setText("Processing...")

        QApplication.processEvents()

        # request_response = {
        #     "data": "ok",
        #     "dsklist": [
        #         {
        #             "dsknumber": "1234",
        #             "box": "07T2406020000223"
        #         },
        #         {
        #             "dsknumber": "5678",
        #             "box": "H632406030000026"
        #         },
        #         {
        #             "dsknumber": "24PLWASW0017705",
        #             "box": "DT32408070695559"
        #         },
        #         {
        #             "dsknumber": "24PLWASW0017599",
        #             "box": "DT32408070629013"
        #         },
        #         {
        #             "dsknumber": "24PLWASW0017563",
        #             "box": "DT32408070606639"
        #         },
        #         {
        #             "dsknumber": "24PLWASW0017562",
        #             "box": "DT32408070606034"
        #         },
        #         {
        #             "dsknumber": "24PLWASW0017565",
        #             "box": "DT32408070606768"
        #         },
        #         {
        #             "dsknumber": "24PLWASW0017564",
        #             "box": "DT32408070606694"
        #         }
        #     ],
        #     "status": "success"
        # }

        # self.json_response = request_response.json()
        try:
            request_response = self.api.get_dsk()
            try:
                if request_response.status_code == 200:
                    self.response_json = request_response.json()
                    print("get response successfully.")
                    if self.response_json.get('status') == 'success':
                        print("get DSK successfully.")
                        QMessageBox.information(self, "Success", "get DSK successfully!")
                        # self.finished = True
                        self.save_button.setEnabled(True)
                        # self.finished_successful.emit()  # 发送信号，通知父窗口更新数据
                    else:
                        err_msg = self.response_json.get('error', 'Unknown error')
                        print(f"get DSK failed: {err_msg}")
                        QMessageBox.critical(self, "Error", f"get DSK failed! {err_msg}")
                else:
                    print(f"Failed to get DSK. Status code: {request_response.status_code}")
                    print("Raw response text:", request_response.text)
                    QMessageBox.critical(self, "Error",
                                         f"Failed to get DSK. Status code: {request_response.status_code}")
            # self.finished = True
            # self.save_button.setEnabled(True)
            # self.finished_successful.emit()
            except AttributeError:
                print("Error: Response object does not have a status_code attribute.")
                QMessageBox.critical(self, "Error", "Invalid response object received.")

        finally:
            # 恢复按钮状态
            self.findChild(QPushButton, "confirmbutton").setEnabled(True)
            self.findChild(QPushButton, "confirmbutton").setText("Submit")
            QApplication.processEvents()

    def on_cancel(self):
        # 创建提示框，确认取消操作
        reply = QMessageBox.question(self, 'Confirm Cancel',
                                     "This action will discard all changes and return to the first page. Are you sure you want to continue?",
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

        if reply == QMessageBox.Yes:
            self.cancel_requested.emit()  # 发送信号，通知父窗口取消操作
        else:
            print("User canceled the action.")

    def save(self):
        if not self.response_json:
            QMessageBox.warning(self, "Response Error", "Response is empty.")
            return
        self.add_dsknumber_to_manifest_with_dialog(global_state.manifest_file_path, self.response_json)

    def add_dsknumber_to_manifest_with_dialog(parent_widget, manifest_path, response_json):
        # Step 1: Parse dsklist to dict
        try:
            dsk_list = response_json.get('dsklist', [])
            if not isinstance(dsk_list, list):
                raise ValueError("Invalid 'dsklist' format: not a list.")
            dsk_dict = {item['box']: item['dsknumber'] for item in dsk_list if 'box' in item and 'dsknumber' in item}
        except Exception as e:
            QMessageBox.critical(parent_widget, "JSON Error", f"Failed to parse DSK data:\n{e}")
            return

        # Step 2: Load Excel
        try:
            wb = load_workbook(manifest_path)
            ws = wb.active
        except Exception as e:
            QMessageBox.critical(parent_widget, "Read Error", f"Failed to open Excel file:\n{e}")
            return

        # Step 3: Find BOXNumber column
        headers = [cell.value for cell in ws[1]]
        try:
            box_col_idx = headers.index('BOXNumber') + 1
        except ValueError:
            QMessageBox.warning(parent_widget, "Missing Column", "The 'BOXNumber' column is missing in the Excel file.")
            return

        # Step 4: Add DSKNumber header
        dsk_col_idx = box_col_idx + 1
        ws.insert_cols(dsk_col_idx)
        ws.cell(row=1, column=dsk_col_idx, value='DSKNumber')

        # Step 5 (before filling): Check if all BOXNumber are matched
        unmatched_boxes = set()

        for row in range(2, ws.max_row + 1):
            box_value = ws.cell(row=row, column=box_col_idx).value
            if box_value not in dsk_dict:
                unmatched_boxes.add(box_value)
            else:
                ws.cell(row=row, column=dsk_col_idx, value=dsk_dict[box_value])

        if unmatched_boxes:
            limited = list(unmatched_boxes)[:10]
            QMessageBox.critical(
                parent_widget,
                "DSK Mapping Error",
                f"The following BOXNumber(s) were not found in the response:\n{', '.join(limited)}" +
                ("\n..." if len(unmatched_boxes) > 10 else "")
            )
            return  # Stop processing if any unmatched

        # Step 5: Fill DSKNumber values
        match_count = 0
        for row in range(2, ws.max_row + 1):
            box_value = ws.cell(row=row, column=box_col_idx).value
            dsk_value = dsk_dict.get(box_value, '')
            if dsk_value:
                match_count += 1
            ws.cell(row=row, column=dsk_col_idx, value=dsk_value)

        if match_count == 0:
            QMessageBox.warning(parent_widget, "No Match",
                                "No matching BOXNumber entries were found in the JSON response.")

        # Step 6: Ask user where to save
        default_filename = f"{global_state.flight}_{global_state.carrier}_dsknumber.xlsx"
        save_path, _ = QFileDialog.getSaveFileName(
            parent_widget,
            "Save Updated Manifest",
            default_filename,
            "Excel Files (*.xlsx)"
        )

        if not save_path:
            return  # User cancelled

        if not save_path.endswith('.xlsx'):
            save_path += '.xlsx'

        # Step 7: Save new file
        try:
            wb.save(save_path)
            QMessageBox.information(parent_widget, "Success", f"File successfully saved to:\n{save_path}")
        except Exception as e:
            QMessageBox.critical(parent_widget, "Save Error", f"Failed to save updated Excel:\n{e}")

        # # Step 1: Parse dsklist to dict
        # dsk_dict = {item['box']: item['dsknumber'] for item in json_response.get('dsklist', [])}
        #
        # # Step 2: Read Excel
        # try:
        #     df = pd.read_excel(manifest_path, dtype=str)
        # except Exception as e:
        #     QMessageBox.critical(parent_widget, "Read Error", f"Failed to read manifest Excel file:\n{e}")
        #     return
        #
        # if 'BOXNumber' not in df.columns:
        #     QMessageBox.warning(parent_widget, "Missing Column", "The 'BOXNumber' column is missing in the Excel file.")
        #     return
        #
        # # Step 3: Map and add new column
        # df['DSKNumber'] = df['BOXNumber'].map(dsk_dict).fillna('')
        # default_filename = f"{global_state.flight}_{global_state.carrier}_dsknumber.xlsx"
        #
        # # Step 4: Let user choose save path
        # save_path, _ = QFileDialog.getSaveFileName(
        #     parent_widget,
        #     "Save Updated Manifest",
        #     default_filename,
        #     "Excel Files (*.xlsx)"
        # )
        #
        # if not save_path:
        #     return  # User cancelled
        #
        # if not save_path.endswith('.xlsx'):
        #     save_path += '.xlsx'
        #
        # # Step 5: Save
        # try:
        #     df.to_excel(save_path, engine='openpyxl', index=False)
        #     QMessageBox.information(parent_widget, "Success", f"File successfully saved to:\n{save_path}")
        # except Exception as e:
        #     QMessageBox.critical(parent_widget, "Save Error", f"Failed to save updated Excel:\n{e}")
