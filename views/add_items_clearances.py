import os

from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtGui import QFont
from PyQt5.QtWidgets import QWidget, QVBoxLayout, QLabel, QLineEdit, QSizePolicy, QPushButton, QHBoxLayout, QFileDialog, \
    QMessageBox, QScrollArea, QApplication, QButtonGroup, QDialog, QListWidget, QSpacerItem, QListWidgetItem, QFrame, \
    QAbstractItemView
import pandas as pd
from openpyxl.reader.excel import load_workbook

import global_state
from http_client import add_items_clearances_request, init_items_clearances_request
from utils.parse_zc415_xml import parse_zc415_xml
from utils.parse_zc428_xml import parse_zc428_xml

RESULTS_DIR = os.path.join(os.getcwd(), "results")
os.makedirs(RESULTS_DIR, exist_ok=True)


# === 模式选择对话框 ===
class ClearanceModeDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Select Processing Mode")
        self.setMinimumWidth(350)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(40, 30, 40, 30)  # 添加左右和上下内边距
        layout.setSpacing(20)

        label = QLabel("Please choose the processing mode:")
        label.setObjectName("normal")
        label.setAlignment(Qt.AlignCenter)
        layout.addWidget(label)

        self.more_than_150_btn = QPushButton(">150")
        self.more_than_150_btn.setObjectName("confirmbutton")
        self.more_than_150_btn.clicked.connect(lambda: self.accept_with_mode(">150"))
        layout.addWidget(self.more_than_150_btn)

        self.less_than_equal_150_btn = QPushButton("<=150")
        self.less_than_equal_150_btn.setObjectName("confirmbutton")
        self.less_than_equal_150_btn.clicked.connect(lambda: self.accept_with_mode("<=150"))
        layout.addWidget(self.less_than_equal_150_btn)

        layout.addItem(QSpacerItem(20, 10))

        self.selected_mode = None

    def accept_with_mode(self, mode):
        self.selected_mode = mode
        self.accept()


# === 自动/手动 模式弹窗（仅用于 <=150） ===
class AutoManualChoiceDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Select ≤150 Mode")
        self.setMinimumWidth(350)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(40, 30, 40, 30)
        layout.setSpacing(20)

        label = QLabel("Choose mode for ≤150:")
        label.setAlignment(Qt.AlignCenter)
        layout.addWidget(label)

        self.auto_btn = QPushButton("Auto Mode (upload process file)")
        self.auto_btn.setObjectName("confirmbutton")
        self.auto_btn.clicked.connect(lambda: self.accept_with_mode("auto"))
        layout.addWidget(self.auto_btn)

        self.manual_btn = QPushButton("Manual Mode (upload ZC415/428)")
        self.manual_btn.setObjectName("confirmbutton")
        self.manual_btn.clicked.connect(lambda: self.accept_with_mode("manual"))
        layout.addWidget(self.manual_btn)

        layout.addItem(QSpacerItem(20, 10))

        self.selected_mode = None

    def accept_with_mode(self, mode):
        self.selected_mode = mode
        self.accept()


# === 多文件上传弹窗（用于 >150 和 手动模式） ===
class MultiFileUploadDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Upload ZC415 and ZC428 Files")
        self.setMinimumWidth(600)

        self.zc415_files = []
        self.zc428_files = []

        layout = QVBoxLayout(self)
        layout.setContentsMargins(40, 30, 40, 30)
        layout.setSpacing(20)

        self.zc415_list = QListWidget()
        btn_415 = QPushButton("Upload ZC415 File")
        btn_415.setFont(QFont("Microsoft YaHei", 12))
        btn_415.setMaximumWidth(800)
        btn_415.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        btn_415.setStyleSheet("padding: 10px;")
        btn_415.clicked.connect(lambda: self.upload_file(self.zc415_list, self.zc415_files))

        self.zc428_list = QListWidget()
        btn_428 = QPushButton("Upload ZC428 File")
        btn_428.setFont(QFont("Microsoft YaHei", 12))
        btn_428.setMaximumWidth(800)
        btn_428.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        btn_428.setStyleSheet("padding: 10px;")
        btn_428.clicked.connect(lambda: self.upload_file(self.zc428_list, self.zc428_files))

        self.get_mrn_btn = QPushButton("Get Clearances MRN")
        self.get_mrn_btn.setObjectName("confirmbutton")
        self.get_mrn_btn.clicked.connect(self.get_mrn)

        lab_415 = QLabel("ZC415 Files:")
        lab_415.setObjectName("normal")
        lab_428 = QLabel("ZC428 Files:")
        lab_428.setObjectName("normal")

        layout.addWidget(lab_415)
        layout.addWidget(self.zc415_list)
        layout.addWidget(btn_415)

        layout.addWidget(lab_428)
        layout.addWidget(self.zc428_list)
        layout.addWidget(btn_428)

        layout.addWidget(self.get_mrn_btn)

    def upload_file(self, list_widget, storage_list):
        files, _ = QFileDialog.getOpenFileNames(self, "Select Files", "", "XML Files (*.xml)")
        for f in files:
            if f not in storage_list:
                storage_list.append(f)
                list_widget.addItem(f)

    def get_mrn(self):
        if not self.zc415_files or not self.zc428_files:
            QMessageBox.warning(self, "Warning", "Please upload both ZC415 and ZC428 files.")
            return

        # === Step1: 解析所有 ZC415、ZC428 ===
        lrn_to_ucr = {}
        for f415 in self.zc415_files:
            with open(f415, "r", encoding="utf-8") as f:
                lrn_to_ucr.update(parse_zc415_xml(f))

        lrn_to_mrn = {}
        for f428 in self.zc428_files:
            with open(f428, "r", encoding="utf-8") as f:
                lrn_to_mrn.update(parse_zc428_xml(f))

        # === Step2: 生成 ucr → mrn 映射 ===
        ucr_to_mrn = {}
        for lrn, ucr in lrn_to_ucr.items():
            mrn = lrn_to_mrn.get(lrn)
            if mrn:
                ucr_to_mrn[ucr] = mrn

        if not ucr_to_mrn:
            QMessageBox.warning(self, "Warning", "No matching LRN found between ZC415 and ZC428.")
            return

        # === Step3: 更新的 Excel 文件 ===

        df = pd.read_excel(global_state.process_file_path, engine="openpyxl")

        # 确保有 MRN 列
        if "MRN(CLEARANCES)" not in df.columns:
            df["MRN(CLEARANCES)"] = ""
        else:
            df["MRN(CLEARANCES)"] = df["MRN(CLEARANCES)"].fillna("").astype(str)

        # 遍历 DataFrame，填充 MRN
        for idx, row in df.iterrows():
            tracking = str(row.get("TrackingNumber", "")).strip()
            if tracking in ucr_to_mrn:
                df.at[idx, "MRN(CLEARANCES)"] = ucr_to_mrn[tracking]

        # 保存结果
        df.to_excel(global_state.process_file_path, index=False)
        QMessageBox.information(self, "Success", "Get Clearance MRN Successfully. Please manually complete the "
                                                 "missing part.")
        self.accept()

        # QMessageBox.information(self, "Get Clearance MRN", "Not Support yet.")
        # self.accept()


class AutoFileUploadDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Upload Manifest File")
        self.setMinimumWidth(500)

        self.file_path = None

        layout = QVBoxLayout(self)
        layout.setContentsMargins(40, 30, 40, 30)
        layout.setSpacing(20)

        self.get_button = QPushButton("Get Clearances MRN")
        self.get_button.setObjectName("confirmbutton")
        self.get_button.clicked.connect(self.get_mrn)
        layout.addWidget(self.get_button)

    def get_mrn(self):
        # 禁用按钮，改样式、文字
        self.findChild(QPushButton, "confirmbutton").setEnabled(False)
        self.findChild(QPushButton, "confirmbutton").setText("Processing...")

        QApplication.processEvents()

        try:
            if not global_state.manifest_file_path:
                QMessageBox.warning(self, "Warning", "Manifest file can't be found.")
                return

            if not global_state.process_file_path:
                QMessageBox.warning(self, "Warning", "Process file can't be found.")

            request_response = self.api.init_items_clearances()

            try:
                response_json = request_response.json()
                if request_response.status_code == 200:
                    print("get response successfully.")
                    self.fill_clearance_column(global_state.process_file_path, response_json)
                    QMessageBox.information(self, "Success", "Get clearances MRN successfully. Please manually "
                                                             "complete the missing part.")
                else:
                    print(f"Failed to init clearance. Status code: {request_response.status_code}")
                    QMessageBox.critical(self, "Error",
                                         f"Failed to init clearance. Status code: {request_response.status_code}")
            except AttributeError:
                print("Error: Response object does not have a status_code attribute.")
                QMessageBox.critical(self, "Error", "Invalid response object received.")

        finally:
            # 恢复按钮状态
            self.findChild(QPushButton, "confirmbutton").setEnabled(True)
            self.findChild(QPushButton, "confirmbutton").setText("Get Clearances MRN")
            QApplication.processEvents()

    def fill_clearance_column(self, process_file_path, response_json):
        # 读取 Excel
        df = pd.read_excel(process_file_path, engine='openpyxl')

        # 避免缺列
        if 'MRN(CLEARANCES)' not in df.columns:
            df['MRN(CLEARANCES)'] = ''
        else:
            df['MRN(CLEARANCES)'] = df['MRN(CLEARANCES)'].fillna('').astype(str)

        # 构造一个清关 MRN 的映射字典 {(box, tracking): clearance}
        mapping = {}
        for entry in response_json:
            box = entry["num"]
            for item in entry["items"]:
                parcel = item["parcel"]
                clearance = item["clereance"]
                mapping[(box, parcel)] = clearance

        # 遍历 DataFrame，填充清关列
        for idx, row in df.iterrows():
            key = (str(row['BOXNumber']), str(row['TrackingNumber']))
            if key in mapping:
                df.at[idx, 'MRN(CLEARANCES)'] = mapping[key]

        df.to_excel(process_file_path, index=False)


class AddItemsClearances(QWidget):
    # finished_successful = pyqtSignal()
    cancel_requested = pyqtSignal()

    def __init__(self, api):

        super().__init__()
        self.column_input = None
        self.step_config = None
        self.api = api
        # self.finished = False

        self.clearance_data = []  # [{ process_file: str, zc415_files: [], zc428_files: []}]

        self.outer_layout = None
        self.scroll_area = None
        self.inner_widget = None
        self.inner_layout = None
        self.scroll_area_label = None
        self.label_status = None

        self.process_list = None
        self.zc415_list = None
        self.zc428_list = None

        self.label_mapping = None
        self.scroll_area_mapping = None

        self.init_ui()

    def set_step_config(self, step_config: dict):
        """设置字段名和提示"""
        self.step_config = step_config
        self.update_hint_label()

    def update_hint_label(self):
        """在页面标题下方增加提示"""
        if not self.step_config or not hasattr(self, "inner_layout"):
            return

        # 删除已有提示标签（如果存在）
        for i in reversed(range(self.inner_layout.count())):
            widget = self.inner_layout.itemAt(i).widget()
            if isinstance(widget, QLabel) and getattr(widget, "is_hint_label", False):
                self.inner_layout.removeWidget(widget)
                widget.deleteLater()

        hint_text = self.step_config.get("hint", "")
        if hint_text:
            hint_label = QLabel(hint_text)
            hint_label.setObjectName("hint")
            hint_label.setAlignment(Qt.AlignCenter)
            hint_label.is_hint_label = True  # 标记为提示
            # 找到标题位置后插入到其下方
            # 假设标题总是 inner_layout 的第一个 QLabel
            for i in range(self.inner_layout.count()):
                widget = self.inner_layout.itemAt(i).widget()
                if isinstance(widget, QLabel) and widget.objectName() == "title":
                    self.inner_layout.insertWidget(i + 1, hint_label)
                    break

    def init_ui(self):

        self.build_outer_layout()

        self.build_scroll_area()

        self.build_inner_layout()

        self.build_title()

        self.build_column_input()

        self.build_process_clearance_panel()

        self.build_action_buttons()

    def reset(self):
        pass

    def build_outer_layout(self):

        # 外层布局（用于居中）
        self.outer_layout = QVBoxLayout(self)
        self.outer_layout.setContentsMargins(0, 0, 0, 0)
        self.outer_layout.setAlignment(Qt.AlignCenter)

    def build_scroll_area(self):
        # 创建 QScrollArea，并设置属性
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)  # 让 scroll area 根据内容自适应
        # self.scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)  # 横向不要滚动条（可选）
        self.scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)  # 纵向滚动条按需出现

        self.outer_layout.addWidget(self.scroll_area)

    def build_inner_layout(self):
        # 内层容器和布局（承载内容）
        self.inner_widget = QWidget()
        self.inner_layout = QVBoxLayout(self.inner_widget)
        self.inner_layout.setContentsMargins(100, 50, 100, 50)
        self.inner_layout.setSpacing(20)
        self.inner_layout.setAlignment(Qt.AlignTop | Qt.AlignHCenter)  # 控制内容居中靠上

        self.scroll_area.setWidget(self.inner_widget)

    def build_title(self):
        # 标题
        title = QLabel("AddItemsClearance")
        title.setObjectName("title")
        title.setAlignment(Qt.AlignCenter)
        self.inner_layout.addWidget(title)
        # self.inner_layout.addSpacing(50)

    def build_column_input(self):
        column_layout = QVBoxLayout()

        column_label = QLabel("Column Number:")
        column_label.setObjectName("normal")
        self.column_input = QLineEdit()
        self.column_input.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        column_layout.addWidget(column_label)
        column_layout.addWidget(self.column_input)
        column_layout.setAlignment(Qt.AlignHCenter)  # 整个控件居中

        self.inner_layout.addLayout(column_layout)
        self.inner_layout.addSpacing(10)

    def build_process_clearance_panel(self):

        process_clearance_layout = QHBoxLayout()

        # 左侧 arn 列表
        left_layout = QVBoxLayout()
        process_label = QLabel("Process Files:")
        process_label.setObjectName("normal")

        self.process_list = QListWidget()
        self.process_list.setSelectionMode(QAbstractItemView.SingleSelection)

        self.process_list.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.process_list.setMinimumHeight(300)

        # self.process_list.setSelectionBehavior(QAbstractItemView.SelectRows)
        # self.process_list.setEditTriggers(QAbstractItemView.NoEditTriggers)
        # self.process_list.horizontalHeader().setStretchLastSection(True)
        # self.process_list.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # self.process_list.verticalHeader().setVisible(False)
        # self.process_list.setSelectionMode(QAbstractItemView.SingleSelection)  # 单选模式

        # self.process_list.setStyleSheet("""
        #     QTableWidget {
        #         background-color: #ffffff;
        #         gridline-color: #ccc;
        #         border: 1px solid #aaa;
        #     }
        #     QHeaderView::section {
        #         background-color: #0078d7;
        #         color: white;
        #         padding: 6px 12px;
        #         border: 1px solid #aaa;
        #     }
        #     QTableWidget::item:selected {
        #         background-color: #87cefa;
        #         color: black;
        #     }
        # """)

        self.process_list.setStyleSheet("""
            QListWidget {
                background-color: #f9f9f9;
                border: 1px solid #ccc;
                padding: 5px;
            }
            QListWidget::item {
                padding: 5px;
            }
            QListWidget::item:selected {
                background-color: #87cefa;
                color: black;
            }
        """)

        self.process_list.itemClicked.connect(self.show_mapping)

        # TODO:按钮设置样式
        process_btn_layout = QHBoxLayout()

        select_all_btn = QPushButton("Select All / Unselect All")
        select_all_btn.setObjectName("fcAddButton")
        select_all_btn.clicked.connect(self.toggle_selects)

        add_process_btn = QPushButton("Add Process File")
        add_process_btn.setObjectName("fcAddButton")
        add_process_btn.clicked.connect(self.add_process_file)

        del_process_btn = QPushButton("Delete Process File")
        del_process_btn.setObjectName("fcDelButton")
        del_process_btn.clicked.connect(self.delete_process_file)

        process_btn_layout.addWidget(select_all_btn)
        process_btn_layout.addWidget(add_process_btn)
        process_btn_layout.addWidget(del_process_btn)

        left_layout.addWidget(process_label)
        left_layout.addWidget(self.process_list)
        left_layout.addLayout(process_btn_layout)

        # 右侧 mapping 文件列表
        right_layout = QVBoxLayout()

        mapping_label = QLabel("ZC415/428 Files:")
        mapping_label.setObjectName("normal")

        # ================= 415 区域 =================
        zc415_layout = QVBoxLayout()

        self.zc415_list = QListWidget()
        self.zc415_list.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.zc415_list.setMinimumHeight(300)
        self.zc415_list.setSelectionMode(QAbstractItemView.NoSelection)  # 单选模式

        self.zc415_list.setStyleSheet("""
            QListWidget {
                background-color: #f9f9f9;
                border: 1px solid #ccc;
                padding: 5px;
            }
            QListWidget::item {
                padding: 5px;
            }
            QListWidget::item:selected {
                background-color: #87cefa;
                color: black;
            }
        """)

        # mapping 操作按钮
        map_btn_layout_415 = QHBoxLayout()

        add_415_btn = QPushButton("Add ZC415 Files")
        add_415_btn.clicked.connect(self.add_415_file)

        del_415_btn = QPushButton("Delete ZC415 Files")
        del_415_btn.clicked.connect(self.delete_415_file)

        add_415_btn.setObjectName("fcAddButton")
        del_415_btn.setObjectName("fcDelButton")

        map_btn_layout_415.addWidget(add_415_btn)
        map_btn_layout_415.addWidget(del_415_btn)

        zc415_label = QLabel("ZC415:")
        zc415_label.setObjectName("normal")
        zc415_layout.addWidget(zc415_label)
        zc415_layout.addWidget(self.zc415_list)
        zc415_layout.addLayout(map_btn_layout_415)

        # ================= 428 区域 =================
        zc428_layout = QVBoxLayout()

        self.zc428_list = QListWidget()
        self.zc428_list.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.zc428_list.setMinimumHeight(300)
        self.zc428_list.setSelectionMode(QAbstractItemView.NoSelection)

        self.zc428_list.setStyleSheet("""
            QListWidget {
                background-color: #f9f9f9;
                border: 1px solid #ccc;
                padding: 5px;
            }
            QListWidget::item {
                padding: 5px;
            }
            QListWidget::item:selected {
                background-color: #87cefa;
                color: black;
            }
        """)

        map_btn_layout_428 = QHBoxLayout()

        add_428_btn = QPushButton("Add ZC428 Files")
        add_428_btn.clicked.connect(self.add_zc428_file)

        del_428_btn = QPushButton("Delete ZC428 Files")
        del_428_btn.clicked.connect(self.delete_zc428_file)

        add_428_btn.setObjectName("fcAddButton")
        del_428_btn.setObjectName("fcDelButton")

        map_btn_layout_428.addWidget(add_428_btn)
        map_btn_layout_428.addWidget(del_428_btn)

        zc428_label = QLabel("ZC428:")
        zc428_label.setObjectName("normal")
        zc428_layout.addWidget(zc428_label)
        zc428_layout.addWidget(self.zc428_list)
        zc428_layout.addLayout(map_btn_layout_428)

        files_layout = QHBoxLayout()
        files_layout.addLayout(zc415_layout)
        files_layout.addLayout(zc428_layout)

        right_layout.addWidget(mapping_label)
        right_layout.addLayout(files_layout)

        process_ens_group = QFrame()
        process_ens_group.setObjectName("fcFrame")
        process_ens_group.setFrameShape(QFrame.NoFrame)
        process_ens_group.setLayout(process_clearance_layout)

        # fc_manifest_layout.addStretch()
        process_clearance_layout.addLayout(left_layout)
        process_clearance_layout.addSpacing(50)
        process_clearance_layout.addLayout(right_layout)
        # fc_manifest_layout.addStretch()

        process_clearance_layout.setStretchFactor(left_layout, 2)
        process_clearance_layout.setStretchFactor(right_layout, 3)

        # self.inner_layout.addLayout(fc_manifest_layout)
        self.inner_layout.addWidget(process_ens_group)

        self.inner_layout.addSpacing(20)

    def build_action_buttons(self):
        write_btn = QPushButton("Write Clearance to Process File")
        write_btn.setFont(QFont("Microsoft YaHei", 12))
        write_btn.setStyleSheet("padding: 10px;")
        write_btn.clicked.connect(self.write_clearance)

        self.inner_layout.addWidget(write_btn)
        self.inner_layout.addSpacing(20)

        get_mrn_button = QPushButton("Get Clearance MRN")
        get_mrn_button.setFont(QFont("Microsoft YaHei", 12))
        get_mrn_button.setStyleSheet("padding: 10px;")
        get_mrn_button.clicked.connect(self.get_clearance_mrn)

        self.inner_layout.addWidget(get_mrn_button)
        self.inner_layout.addSpacing(20)

        btn_layout = QHBoxLayout()

        submit_btn = QPushButton("Submit")
        submit_btn.setObjectName("confirmbutton")
        submit_btn.clicked.connect(self.submit)

        cancel_btn = QPushButton("Cancel")
        cancel_btn.setObjectName("cancelButton")
        cancel_btn.clicked.connect(self.on_cancel)

        btn_layout.addWidget(submit_btn)
        btn_layout.addWidget(cancel_btn)

        self.inner_layout.addLayout(btn_layout)
        self.inner_layout.addSpacing(20)
        self.inner_layout.addStretch()

    def show_mapping(self, item):
        self.zc415_list.clear()
        self.zc428_list.clear()

        row = self.process_list.row(item)  # ✅ QTableWidgetItem → int
        if row < 0 or row >= len(self.clearance_data):
            return

        data = self.clearance_data[row]

        # 显示 415
        for f in data.get("zc415_files", []):
            list_item = QListWidgetItem(f)
            list_item.setToolTip(f)
            self.zc415_list.addItem(list_item)

        # 显示 428
        for f in data.get("zc428_files", []):
            list_item = QListWidgetItem(f)
            list_item.setToolTip(f)
            self.zc428_list.addItem(list_item)

    def toggle_selects(self):
        if self.process_list.count() == 0:
            return

        # 判断是否已经全部选中
        all_checked = True
        for i in range(self.process_list.count()):
            if self.process_list.item(i).checkState() != Qt.Checked:
                all_checked = False
                break

        # 如果全部选中 → 取消
        new_state = Qt.Unchecked if all_checked else Qt.Checked

        for i in range(self.process_list.count()):
            self.process_list.item(i).setCheckState(new_state)

    def add_process_file(self):
        file_dialog = QFileDialog(self)
        file_dialog.setWindowTitle("Select Process Files")
        file_dialog.setFileMode(QFileDialog.ExistingFiles)
        file_dialog.setNameFilter("Excel Files (*.xls *.xlsx)")

        if file_dialog.exec_():

            file_paths = file_dialog.selectedFiles()

            for file_path in file_paths:

                if any(d["process_file"] == file_path for d in self.clearance_data):
                    continue  # 已存在就跳过

                self.clearance_data.append({
                    "process_file": file_path,
                    "zc415_files": [],
                    "zc428_files": []
                })

                list_item = QListWidgetItem(file_path)
                list_item.setToolTip(file_path)  # 鼠标悬浮显示完整路径

                # 加 checkbox（和 ENS 一样）
                list_item.setFlags(list_item.flags() | Qt.ItemIsUserCheckable)
                list_item.setCheckState(Qt.Unchecked)

                self.process_list.addItem(list_item)

    def delete_process_file(self):
        rows_to_delete = []

        for row in range(self.process_list.count()):
            item = self.process_list.item(row)

            if item.checkState() == Qt.Checked:
                rows_to_delete.append(row)

        if not rows_to_delete:
            QMessageBox.warning(self, "Warning", "Please select process files to delete.")
            return

        for row in reversed(rows_to_delete):
            self.process_list.takeItem(row)
            del self.clearance_data[row]

        self.zc415_list.clear()
        self.zc428_list.clear()

    def add_415_file(self):
        row = self.process_list.currentRow()

        if row < 0:
            QMessageBox.warning(self, "Warning", "Please select a process file first.")
            return

        file_dialog = QFileDialog(self)
        file_dialog.setWindowTitle("Select ZC415 Files")
        file_dialog.setFileMode(QFileDialog.ExistingFiles)
        file_dialog.setNameFilter("XML Files (*.xml)")

        if file_dialog.exec_():
            self.zc415_list.clear()
            self.clearance_data[row]["zc415_files"] = []

            for file_path in file_dialog.selectedFiles():
                self.clearance_data[row]["zc415_files"].append(file_path)

                list_item = QListWidgetItem(file_path)
                list_item.setToolTip(file_path)  # 鼠标悬浮显示完整路径
                self.zc415_list.addItem(list_item)

    def delete_415_file(self):
        row = self.process_list.currentRow()

        if row < 0:
            return

        self.clearance_data[row]["zc415_files"].clear()

        self.zc415_list.clear()

    def add_zc428_file(self):
        row = self.process_list.currentRow()

        if row < 0:
            QMessageBox.warning(self, "Warning", "Please select a process file first.")
            return

        file_dialog = QFileDialog(self)
        file_dialog.setWindowTitle("Select ZC428 Files")
        file_dialog.setFileMode(QFileDialog.ExistingFiles)
        file_dialog.setNameFilter("XML Files (*.xml)")

        if file_dialog.exec_():
            self.zc428_list.clear()
            self.clearance_data[row]["zc428_files"] = []

            for file_path in file_dialog.selectedFiles():
                self.clearance_data[row]["zc428_files"].append(file_path)

                list_item = QListWidgetItem(file_path)
                list_item.setToolTip(file_path)  # 鼠标悬浮显示完整路径
                self.zc428_list.addItem(list_item)

    def delete_zc428_file(self):
        row = self.process_list.currentRow()

        if row < 0:
            return

        self.clearance_data[row]["zc428_files"].clear()

        self.zc428_list.clear()

    def write_clearance(self):
        if not self.clearance_data:
            QMessageBox.warning(self, "Warning", "No process file to write.")
            return

            # 检查是否有缺少 mapping 的
        missing = []
        for data in self.clearance_data:
            if not data["zc415_files"] or not data["zc428_files"]:
                missing.append(data["process_file"])

        if missing:
            msg = "The following process files are missing ZC415 or ZC428 files:\n\n"
            msg += "\n".join(missing)
            QMessageBox.warning(self, "Missing XML Files", msg)
            return

        # 全部都有 mapping → 开始写 ENS
        success_list = []
        failed_list = []

        for data in self.clearance_data:

            process_path = data["process_file"]
            zc415_files = data["zc415_files"]
            zc428_files = data["zc428_files"]

            try:
                self.write_single_process_clearance(
                    process_path,
                    zc415_files,
                    zc428_files
                )
                success_list.append(process_path)

            except Exception as e:
                failed_list.append(f"{process_path} -> {str(e)}")

        msg = f"Clearance write completed.\n\nSuccess: {len(success_list)} file(s)"

        if failed_list:
            msg += f"\nFailed: {len(failed_list)} file(s)\n" + "\n".join(failed_list)

        QMessageBox.information(self, "Result", msg)

    def write_single_process_clearance(self, process_path, zc415_files, zc428_files):

        # === Step1: 解析 XML ===
        lrn_to_ucr = {}

        for f415 in zc415_files:
            with open(f415, "r", encoding="utf-8") as f:
                lrn_to_ucr.update(parse_zc415_xml(f))

        lrn_to_mrn = {}

        for f428 in zc428_files:
            with open(f428, "r", encoding="utf-8") as f:
                lrn_to_mrn.update(parse_zc428_xml(f))

        # === Step2: 构造 ucr → mrn ===
        ucr_to_mrn = {}

        for lrn, ucr in lrn_to_ucr.items():
            mrn = lrn_to_mrn.get(lrn)
            if mrn:
                ucr_to_mrn[ucr] = mrn

        if not ucr_to_mrn:
            raise ValueError("No matching LRN found between ZC415 and ZC428.")

        # === Step3: 读取 Excel ===
        wb = load_workbook(process_path)
        ws = wb.active

        tracking_col = 2
        mrn_col = 5

        column_text = self.column_input.text().strip()
        if column_text:

            if not column_text.isdigit():
                QMessageBox.warning(self, "Warning", "Please input a valid column number.")
                return
            mrn_col = int(column_text)


        # === Step4: 填充 MRN ===
        for row in range(2, ws.max_row + 1):
            tracking = str(ws.cell(row=row, column=tracking_col).value).strip()
            if tracking in ucr_to_mrn:
                ws.cell(row=row, column=mrn_col).value = ucr_to_mrn[tracking]

        wb.save(process_path)

    def submit(self):
        # 禁用按钮，改样式、文字
        self.findChild(QPushButton, "confirmbutton").setEnabled(False)
        self.findChild(QPushButton, "confirmbutton").setText("Processing...")

        QApplication.processEvents()

        try:
            request_response = self.api.add_items_clearances()
            try:
                if request_response.status_code == 200:
                    response_json = request_response.json()
                    print("get response successfully.")
                    if response_json.get('success', False):
                        print("Add items clearance successfully.")
                        QMessageBox.information(self, "Success", "Add items clearance successfully!")
                        # self.finished = True
                        # self.finished_successful.emit()  # 发送信号，通知父窗口更新数据
                    else:
                        err_msg = response_json.get('error', 'Unknown error')
                        print(f"Add items clearance failed: {err_msg}")
                        QMessageBox.critical(self, "Error", f"Add items clearance failed! {err_msg}")
                else:
                    print(f"Failed to Add items clearance. Status code: {request_response.status_code}")
                    print("Raw response text:", request_response.text)
                    QMessageBox.critical(self, "Error",
                                         f"Failed to Add items clearance. Status code: {request_response.status_code}")

            except AttributeError:
                print("Error: Response object does not have a status_code attribute.")
                QMessageBox.critical(self, "Error", "Invalid response object received.")

        finally:
            # 恢复按钮状态
            self.findChild(QPushButton, "confirmbutton").setEnabled(True)
            self.findChild(QPushButton, "confirmbutton").setText("Submit")
            QApplication.processEvents()

    def on_cancel(self):
        # 创建提示框，确认取消操作
        reply = QMessageBox.question(self, 'Confirm Cancel',
                                     "This action will discard all changes and return to the first page. Are you sure you want to continue?",
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

        if reply == QMessageBox.Yes:
            self.cancel_requested.emit()  # 发送信号，通知父窗口取消操作
        else:
            print("User canceled the action.")

    def select_file(self):

        file_dialog = QFileDialog(self)
        file_dialog.setWindowTitle('Select Process File')
        file_dialog.setFileMode(QFileDialog.ExistingFile)  # 只允许选择一个文件
        file_dialog.setNameFilter("Excel Files (*.xls *.xlsx)")

        if file_dialog.exec_():
            selected_file = file_dialog.selectedFiles()[0]  # 只取第一个
            self.file_path = selected_file  # 保存文件路径
            self.label_status.setText(self.file_path)  # 显示文件路径

    def get_clearance_mrn(self):
        dialog = ClearanceModeDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            mode = dialog.selected_mode
            if mode == '>150':
                MultiFileUploadDialog(self).exec_()
            else:
                choice_dialog = AutoManualChoiceDialog(self)
                if choice_dialog.exec_() == QDialog.Accepted:
                    if choice_dialog.selected_mode == 'auto':
                        AutoFileUploadDialog(self).exec_()
                    else:
                        MultiFileUploadDialog(self).exec_()

    def save(self):
        pass
        # get clearance后，保存文件
