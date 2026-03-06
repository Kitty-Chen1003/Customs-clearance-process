import pandas as pd
from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtGui import QFont
from PyQt5.QtWidgets import QWidget, QVBoxLayout, QLabel, QLineEdit, QSizePolicy, QPushButton, QHBoxLayout, QFileDialog, \
    QMessageBox, QScrollArea, QApplication, QListWidget, QAbstractItemView, QFrame, QListWidgetItem
from openpyxl.reader.excel import load_workbook

import global_state
from http_client import add_ens_request


class AddEns(QWidget):
    # finished_successful = pyqtSignal()
    cancel_requested = pyqtSignal()

    def __init__(self, api):

        super().__init__()
        self.column_input = None
        self.step_config = None
        self.api = api

        self.ens_data = []  # [{ process_file: str, mapping_file: str }]

        self.outer_layout = None
        self.scroll_area = None
        self.inner_widget = None
        self.inner_layout = None
        self.scroll_area_label = None
        self.label_status = None

        self.process_list = None
        self.mapping_list = None

        self.label_mapping = None
        self.scroll_area_mapping = None

        self.init_ui()

    def set_step_config(self, step_config: dict):
        """设置字段名和提示"""
        self.step_config = step_config
        self.update_hint_label()

    def update_hint_label(self):
        """在页面标题下方增加提示"""
        if not self.step_config or not hasattr(self, "inner_layout"):
            return

        # 删除已有提示标签（如果存在）
        for i in reversed(range(self.inner_layout.count())):
            widget = self.inner_layout.itemAt(i).widget()
            if isinstance(widget, QLabel) and getattr(widget, "is_hint_label", False):
                self.inner_layout.removeWidget(widget)
                widget.deleteLater()

        hint_text = self.step_config.get("hint", "")
        if hint_text:
            hint_label = QLabel(hint_text)
            hint_label.setObjectName("hint")
            hint_label.setAlignment(Qt.AlignCenter)
            hint_label.is_hint_label = True  # 标记为提示
            # 找到标题位置后插入到其下方
            # 假设标题总是 inner_layout 的第一个 QLabel
            for i in range(self.inner_layout.count()):
                widget = self.inner_layout.itemAt(i).widget()
                if isinstance(widget, QLabel) and widget.objectName() == "title":
                    self.inner_layout.insertWidget(i + 1, hint_label)
                    break

    def init_ui(self):

        self.build_outer_layout()

        self.build_scroll_area()

        self.build_inner_layout()

        self.build_title()

        self.build_column_input()

        self.build_process_ens_panel()

        self.build_action_buttons()

    def build_outer_layout(self):

        # 外层布局（用于居中）
        self.outer_layout = QVBoxLayout(self)
        self.outer_layout.setContentsMargins(0, 0, 0, 0)
        self.outer_layout.setAlignment(Qt.AlignCenter)

    def build_scroll_area(self):
        # 创建 QScrollArea，并设置属性
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)  # 让 scroll area 根据内容自适应
        # self.scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)  # 横向不要滚动条（可选）
        self.scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)  # 纵向滚动条按需出现

        self.outer_layout.addWidget(self.scroll_area)

    def build_inner_layout(self):
        # 内层容器和布局（承载内容）
        self.inner_widget = QWidget()
        self.inner_layout = QVBoxLayout(self.inner_widget)
        self.inner_layout.setContentsMargins(100, 50, 100, 50)
        self.inner_layout.setSpacing(20)
        self.inner_layout.setAlignment(Qt.AlignTop | Qt.AlignHCenter)  # 控制内容居中靠上

        self.scroll_area.setWidget(self.inner_widget)

    def build_title(self):
        # 标题
        title = QLabel("AddENS")
        title.setObjectName("title")
        title.setAlignment(Qt.AlignCenter)
        self.inner_layout.addWidget(title)
        # self.inner_layout.addSpacing(50)

    def build_column_input(self):
        column_layout = QVBoxLayout()

        column_label = QLabel("Column Number:")
        column_label.setObjectName("normal")
        self.column_input = QLineEdit()
        self.column_input.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        column_layout.addWidget(column_label)
        column_layout.addWidget(self.column_input)
        column_layout.setAlignment(Qt.AlignHCenter)  # 整个控件居中

        self.inner_layout.addLayout(column_layout)
        self.inner_layout.addSpacing(10)

    def build_process_ens_panel(self):

        process_ens_layout = QHBoxLayout()

        # 左侧 arn 列表
        left_layout = QVBoxLayout()
        process_label = QLabel("Process Files:")
        process_label.setObjectName("normal")

        # self.process_list = QTableWidget()
        # self.process_list.setColumnCount(1)
        # self.process_list.setHorizontalHeaderLabels(["ARN"])
        self.process_list = QListWidget()
        self.process_list.setSelectionMode(QAbstractItemView.SingleSelection)

        self.process_list.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.process_list.setMinimumHeight(300)

        # self.process_list.setSelectionBehavior(QAbstractItemView.SelectRows)
        # self.process_list.setEditTriggers(QAbstractItemView.NoEditTriggers)
        # self.process_list.horizontalHeader().setStretchLastSection(True)
        # self.process_list.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # self.process_list.verticalHeader().setVisible(False)
        # self.process_list.setSelectionMode(QAbstractItemView.SingleSelection)  # 单选模式

        # self.process_list.setStyleSheet("""
        #     QTableWidget {
        #         background-color: #ffffff;
        #         gridline-color: #ccc;
        #         border: 1px solid #aaa;
        #     }
        #     QHeaderView::section {
        #         background-color: #0078d7;
        #         color: white;
        #         padding: 6px 12px;
        #         border: 1px solid #aaa;
        #     }
        #     QTableWidget::item:selected {
        #         background-color: #87cefa;
        #         color: black;
        #     }
        # """)

        self.process_list.setStyleSheet("""
            QListWidget {
                background-color: #f9f9f9;
                border: 1px solid #ccc;
                padding: 5px;
            }
            QListWidget::item {
                padding: 5px;
            }
            QListWidget::item:selected {
                background-color: #87cefa;
                color: black;
            }
        """)

        self.process_list.itemClicked.connect(self.show_mapping)

        # TODO:按钮设置样式
        process_btn_layout = QHBoxLayout()

        select_all_btn = QPushButton("Select All / Unselect All")
        select_all_btn.setObjectName("fcAddButton")
        select_all_btn.clicked.connect(self.toggle_selects)

        add_process_btn = QPushButton("Add Process File")
        add_process_btn.setObjectName("fcAddButton")
        add_process_btn.clicked.connect(self.add_process_file)

        del_process_btn = QPushButton("Delete Process File")
        del_process_btn.setObjectName("fcDelButton")
        del_process_btn.clicked.connect(self.delete_process_file)

        process_btn_layout.addWidget(select_all_btn)
        process_btn_layout.addWidget(add_process_btn)
        process_btn_layout.addWidget(del_process_btn)

        left_layout.addWidget(process_label)
        left_layout.addWidget(self.process_list)
        left_layout.addLayout(process_btn_layout)

        # 右侧 mapping 文件列表
        right_layout = QVBoxLayout()
        mapping_label = QLabel("Mapping File:")
        mapping_label.setObjectName("normal")

        self.mapping_list = QListWidget()
        self.mapping_list.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.mapping_list.setMinimumHeight(300)
        self.mapping_list.setSelectionMode(QAbstractItemView.NoSelection)  # 单选模式

        self.mapping_list.setStyleSheet("""
            QListWidget {
                background-color: #f9f9f9;
                border: 1px solid #ccc;
                padding: 5px;
            }
            QListWidget::item {
                padding: 5px;
            }
            QListWidget::item:selected {
                background-color: #87cefa;
                color: black;
            }
        """)

        # mapping 操作按钮
        map_btn_layout = QHBoxLayout()
        add_map_btn = QPushButton("Add Mapping File")
        add_map_btn.clicked.connect(self.add_mapping_file)

        del_map_btn = QPushButton("Delete Mapping File")
        del_map_btn.clicked.connect(self.delete_mapping_file)

        add_map_btn.setObjectName("fcAddButton")
        del_map_btn.setObjectName("fcDelButton")

        map_btn_layout.addWidget(add_map_btn)
        map_btn_layout.addWidget(del_map_btn)

        right_layout.addWidget(mapping_label)
        right_layout.addWidget(self.mapping_list)
        right_layout.addLayout(map_btn_layout)

        process_ens_group = QFrame()
        process_ens_group.setObjectName("fcFrame")
        process_ens_group.setFrameShape(QFrame.NoFrame)
        process_ens_group.setLayout(process_ens_layout)

        # fc_manifest_layout.addStretch()
        process_ens_layout.addLayout(left_layout)
        process_ens_layout.addSpacing(50)
        process_ens_layout.addLayout(right_layout)
        # fc_manifest_layout.addStretch()

        process_ens_layout.setStretchFactor(left_layout, 1)
        process_ens_layout.setStretchFactor(right_layout, 1)

        # self.inner_layout.addLayout(fc_manifest_layout)
        self.inner_layout.addWidget(process_ens_group)

        self.inner_layout.addSpacing(20)

    # def reset(self):
    #     # 重置文件路径和上传按钮文本
    #     self.file_path = ""
    #     self.label_status.setText("No file selected")

    def build_action_buttons(self):
        write_btn = QPushButton("Write ENS to Process File")
        write_btn.setFont(QFont("Microsoft YaHei", 12))
        write_btn.setStyleSheet("padding: 10px;")
        write_btn.clicked.connect(self.write_ens)

        self.inner_layout.addWidget(write_btn)
        self.inner_layout.addSpacing(20)

        btn_layout = QHBoxLayout()

        submit_btn = QPushButton("Submit")
        submit_btn.setObjectName("confirmbutton")
        submit_btn.clicked.connect(self.submit)

        cancel_btn = QPushButton("Cancel")
        cancel_btn.setObjectName("cancelButton")
        cancel_btn.clicked.connect(self.on_cancel)

        btn_layout.addWidget(submit_btn)
        btn_layout.addWidget(cancel_btn)

        self.inner_layout.addLayout(btn_layout)
        self.inner_layout.addSpacing(20)
        self.inner_layout.addStretch()

    def show_mapping(self, item):
        self.mapping_list.clear()

        row = self.process_list.row(item)  # ✅ QTableWidgetItem → int
        if row < 0 or row >= len(self.ens_data):
            return

        mapping_file = self.ens_data[row].get("mapping_file")

        if mapping_file:
            list_item = QListWidgetItem(mapping_file)
            list_item.setToolTip(mapping_file)
            self.mapping_list.addItem(list_item)

    def toggle_selects(self):
        if self.process_list.count() == 0:
            return

        # 判断是否已经全部选中
        all_checked = True
        for i in range(self.process_list.count()):
            if self.process_list.item(i).checkState() != Qt.Checked:
                all_checked = False
                break

        # 如果全部选中 → 取消
        new_state = Qt.Unchecked if all_checked else Qt.Checked

        for i in range(self.process_list.count()):
            self.process_list.item(i).setCheckState(new_state)

    def add_process_file(self):
        file_dialog = QFileDialog(self)
        file_dialog.setWindowTitle("Select Process Files")
        file_dialog.setFileMode(QFileDialog.ExistingFiles)
        file_dialog.setNameFilter("Excel Files (*.xls *.xlsx)")

        if file_dialog.exec_():

            file_paths = file_dialog.selectedFiles()

            for file_path in file_paths:

                if any(d["process_file"] == file_path for d in self.ens_data):
                    continue  # 已存在就跳过

                self.ens_data.append({
                    "process_file": file_path,
                    "mapping_file": ""
                })

                list_item = QListWidgetItem(file_path)
                list_item.setToolTip(file_path)  # 鼠标悬浮显示完整路径

                # 添加 checkbox
                list_item.setFlags(list_item.flags() | Qt.ItemIsUserCheckable)
                list_item.setCheckState(Qt.Unchecked)

                self.process_list.addItem(list_item)

    def delete_process_file(self):

        rows_to_delete = []

        for row in range(self.process_list.count()):
            item = self.process_list.item(row)

            if item.checkState() == Qt.Checked:
                rows_to_delete.append(row)

        if not rows_to_delete:
            QMessageBox.warning(self, "Warning", "Please select process files to delete.")
            return

        for row in reversed(rows_to_delete):
            self.process_list.takeItem(row)
            del self.ens_data[row]

        self.mapping_list.clear()

    def add_mapping_file(self):
        row = self.process_list.currentRow()

        if row < 0:
            QMessageBox.warning(self, "Warning", "Please select a process file first.")
            return

        file_dialog = QFileDialog(self)
        file_dialog.setWindowTitle("Select Mapping File")
        file_dialog.setFileMode(QFileDialog.ExistingFile)
        file_dialog.setNameFilter("Excel Files (*.xls *.xlsx)")

        if file_dialog.exec_():
            file_path = file_dialog.selectedFiles()[0]

            self.ens_data[row]["mapping_file"] = file_path

            self.mapping_list.clear()
            list_item = QListWidgetItem(file_path)
            list_item.setToolTip(file_path)  # 鼠标悬浮显示完整路径
            self.mapping_list.addItem(list_item)

    def delete_mapping_file(self):
        row = self.process_list.currentRow()

        if row < 0:
            return

        self.ens_data[row]["mapping_file"] = ""

        self.mapping_list.clear()

    def write_ens(self):
        if not self.ens_data:
            QMessageBox.warning(self, "Warning", "No process file to write.")
            return

            # 检查是否有缺少 mapping 的
        missing = [data["process_file"] for data in self.ens_data if not data.get("mapping_file")]

        if missing:
            msg = "The following process files have no mapping file:\n\n"
            msg += "\n".join(missing)

            QMessageBox.warning(self, "Missing Mapping File", msg)
            return

        # 全部都有 mapping → 开始写 ENS
        success_list = []
        failed_list = []

        for data in self.ens_data:

            process_path = data["process_file"]
            mapping_path = data["mapping_file"]
            try:

                self.write_f26_mrn(process_path, mapping_path)
                success_list.append(process_path)

            except Exception as e:
                failed_list.append(f"{process_path} -> {str(e)}")

        # 构造提示信息
        msg = f"ENS write completed.\n\nSuccess: {len(success_list)} file(s)"
        if failed_list:
            msg += f"\nFailed: {len(failed_list)} file(s)\n" + "\n".join(failed_list)

        QMessageBox.information(self, "Result", msg)

    def select_mapping_file(self):

        file_dialog = QFileDialog(self)
        file_dialog.setWindowTitle('Select Mapping File')
        file_dialog.setFileMode(QFileDialog.ExistingFile)  # 只允许选择一个文件
        file_dialog.setNameFilter("Excel File (*.xls *.xlsx)")

        if file_dialog.exec_():
            selected_file = file_dialog.selectedFiles()[0]  # 只取第一个
            self.mapping_path = selected_file  # 保存文件路径
            self.label_mapping.setText(self.mapping_path)  # 显示文件路径

    # === 写入逻辑 ===
    def write_f26_mrn(self, process_path, mapping_path):

        # === 读取 mapping 文件 ===

        # mapping_dict = {
        #     (row['Container code'], row['Tracking code']): row['F26 MRN']
        #     for _, row in mapping_df.iterrows()
        # }
        try:
            mapping_df = pd.read_excel(mapping_path, dtype=str).fillna('')
            mapping_dict = dict(
                zip(
                    zip(mapping_df['Container code'], mapping_df['Tracking code']),
                    mapping_df['F26 MRN']
                )
            )
        except KeyError as e:
            missing_col = e.args[0]  # 会是 'Container code' 或 'Tracking code' 或 'F26 MRN'
            raise ValueError(f"Mapping file {mapping_path} is missing required column: {missing_col}") from e

        # === 打开 process 文件 ===
        wb = load_workbook(process_path)
        ws = wb.active

        box_idx = 1
        tracking_idx = 2
        f26_idx = 4

        column_text = self.column_input.text().strip()
        if column_text:

            if not column_text.isdigit():
                QMessageBox.warning(self, "Warning", "Please input a valid column number.")
                return
            f26_idx = int(column_text)

        # 填充
        for row in range(2, ws.max_row + 1):
            box_val = ws.cell(row=row, column=box_idx).value
            tracking_val = ws.cell(row=row, column=tracking_idx).value
            key = (str(box_val), str(tracking_val))
            ws.cell(row=row, column=f26_idx, value=mapping_dict.get(key, ''))

        # 保存结果
        wb.save(process_path)

    def submit(self):
        if not self.check_process_file_f26():
            return
        # 禁用按钮，改样式、文字
        self.findChild(QPushButton, "confirmbutton").setEnabled(False)
        self.findChild(QPushButton, "confirmbutton").setText("Processing...")

        QApplication.processEvents()

        try:

            request_response = self.api.add_ens()
            # self.finished = True
            # self.finished_successful.emit()
            try:
                if request_response.status_code == 200:
                    response_json = request_response.json()
                    print("get response successfully.")
                    if response_json.get('success', False):
                        print("Add Ens successfully.")
                        QMessageBox.information(self, "Success", "Add Ens successfully!")
                        # self.finished = True
                        # self.finished_successful.emit()  # 发送信号，通知父窗口更新数据
                    else:
                        err_msg = response_json.get('error', 'Unknown error')
                        print(f"Add Ens failed: {err_msg}")
                        QMessageBox.critical(self, "Error", f"Add Ens failed! {err_msg}")

                else:
                    print(f"Failed to add Ens. Status code: {request_response.status_code}")
                    print("Raw response text:", request_response.text)
                    QMessageBox.critical(self, "Error",
                                         f"Failed to add Ens. Status code: {request_response.status_code}")
            except AttributeError:
                print("Error: Response object does not have a status_code attribute.")
                QMessageBox.critical(self, "Error", "Invalid response object received.")

        finally:
            # 恢复按钮状态
            self.findChild(QPushButton, "confirmbutton").setEnabled(True)
            self.findChild(QPushButton, "confirmbutton").setText("Submit")
            QApplication.processEvents()

    def on_cancel(self):
        # 创建提示框，确认取消操作
        reply = QMessageBox.question(self, 'Confirm Cancel',
                                     "This action will discard all changes and return to the first page. Are you sure you want to continue?",
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

        if reply == QMessageBox.Yes:
            self.cancel_requested.emit()  # 发送信号，通知父窗口取消操作
        else:
            print("User canceled the action.")

    # def select_file(self):
    #
    #     file_dialog = QFileDialog(self)
    #     file_dialog.setWindowTitle('Select Process File')
    #     file_dialog.setFileMode(QFileDialog.ExistingFile)  # 只允许选择一个文件
    #     file_dialog.setNameFilter("Excel Files (*.xls *.xlsx)")
    #
    #     if file_dialog.exec_():
    #         selected_file = file_dialog.selectedFiles()[0]  # 只取第一个
    #         self.file_path = selected_file  # 保存文件路径
    #         self.label_status.setText(self.file_path)  # 显示文件路径

    def check_process_file_f26(self):
        try:
            if not global_state.process_file_path:
                QMessageBox.warning(self, "Missing File", "No process file found.")
                return False

            df = pd.read_excel(global_state.process_file_path, dtype=str)
            df = df.fillna('')

            ens_col = df.columns[3]  # 取第4列的列名
            missing_rows = df[df.iloc[:, 3].str.strip() == ""]

            if not missing_rows.empty:
                QMessageBox.critical(
                    self,
                    "Error",
                    f"There are {len(missing_rows)} rows missing MRN(ENS) values.\n "
                    f"Please complete them before submitting."
                )
                return False

            return True

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to check process file:\n{e}")
            return False
