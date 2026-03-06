import os
import time

import openpyxl
import pandas as pd
import requests
from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtGui import QFont
from PyQt5.QtWidgets import QWidget, QVBoxLayout, QLabel, QLineEdit, QSizePolicy, QPushButton, QHBoxLayout, QFileDialog, \
    QMessageBox, QScrollArea, QApplication, QListWidget, QListWidgetItem, QTableWidget, QAbstractItemView, QHeaderView, \
    QTableWidgetItem, QFrame
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import global_state
from http_client import add_awb_request
from utils.utils import normalize_columns


class AddAwb(QWidget):
    cancel_requested = pyqtSignal()

    def __init__(self, api):

        super().__init__()

        self.step_config = None
        self.api = api
        self.flights_data = []  # [{'flight': flight_number, 'manifests': [file1, file2]}]

        self.outer_layout = None
        self.scroll_area = None
        self.inner_widget = None
        self.inner_layout = None

        self.scroll_area_label = None
        self.label_status = None
        self.flight_input = None
        self.carrier_input = None
        self.fc_list = None
        self.manifest_list = None

        self.file_path = ""

        self.init_ui()
        
    def set_step_config(self, step_config: dict):
        """设置字段名和提示"""
        self.step_config = step_config
        self.update_hint_label()

    def update_hint_label(self):
        """在页面标题下方增加提示"""
        if not self.step_config or not hasattr(self, "inner_layout"):
            return

        # 删除已有提示标签（如果存在）
        for i in reversed(range(self.inner_layout.count())):
            widget = self.inner_layout.itemAt(i).widget()
            if isinstance(widget, QLabel) and getattr(widget, "is_hint_label", False):
                self.inner_layout.removeWidget(widget)
                widget.deleteLater()

        hint_text = self.step_config.get("hint", "")
        if hint_text:
            hint_label = QLabel(hint_text)
            hint_label.setObjectName("hint")
            hint_label.setAlignment(Qt.AlignCenter)
            hint_label.is_hint_label = True  # 标记为提示
            # 找到标题位置后插入到其下方
            # 假设标题总是 inner_layout 的第一个 QLabel
            for i in range(self.inner_layout.count()):
                widget = self.inner_layout.itemAt(i).widget()
                if isinstance(widget, QLabel) and widget.objectName() == "title":
                    self.inner_layout.insertWidget(i + 1, hint_label)
                    break

    def init_ui(self):

        self.build_outer_layout()

        self.build_scroll_area()

        self.build_inner_layout()

        self.build_title()

        self.build_flight_carrier_input()

        self.build_fc_manifest_panel()

        self.build_action_buttons()

        # 文件上传按钮
        # upload_label = QLabel("upload manifest file:")
        # upload_label.setObjectName("normal")
        # inner_layout.addWidget(upload_label)

        # upload_button = QPushButton("Select file")
        # upload_button.setFont(QFont("Microsoft YaHei", 12))
        # upload_button.setMaximumWidth(800)
        # upload_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        # upload_button.setStyleSheet("padding: 10px;")
        # upload_button.clicked.connect(self.select_file)

        # 创建 QLabel 并放入 QScrollArea 中
        # self.label_status = QLabel('No file selected')
        # self.label_status.setWordWrap(True)
        # self.label_status.setObjectName("normal")
        # self.scroll_area_label = QScrollArea()
        # self.scroll_area_label.setMaximumWidth(800)
        # self.scroll_area_label.setWidgetResizable(True)
        # self.scroll_area_label.setMinimumHeight(100)
        # self.scroll_area_label.setWidget(self.label_status)

        # inner_layout.addWidget(self.scroll_area_label)
        # inner_layout.addSpacing(20)

    def build_outer_layout(self):

        # 外层布局（用于居中）
        self.outer_layout = QVBoxLayout(self)
        self.outer_layout.setContentsMargins(0, 0, 0, 0)
        self.outer_layout.setAlignment(Qt.AlignCenter)

    def build_scroll_area(self):
        # 创建 QScrollArea，并设置属性
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)  # 让 scroll area 根据内容自适应
        # self.scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)  # 横向不要滚动条（可选）
        self.scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)  # 纵向滚动条按需出现

        self.outer_layout.addWidget(self.scroll_area)

    def build_inner_layout(self):
        # 内层容器和布局（承载内容）
        self.inner_widget = QWidget()
        self.inner_layout = QVBoxLayout(self.inner_widget)
        self.inner_layout.setContentsMargins(100, 50, 100, 50)
        self.inner_layout.setSpacing(20)
        self.inner_layout.setAlignment(Qt.AlignTop | Qt.AlignHCenter)  # 控制内容居中靠上

        self.scroll_area.setWidget(self.inner_widget)

    def build_title(self):
        # 标题
        title = QLabel("addAWB")
        title.setObjectName("title")
        title.setAlignment(Qt.AlignCenter)
        self.inner_layout.addWidget(title)

    def build_flight_carrier_input(self):
        # 输入框
        flight_carrier_layout = QVBoxLayout()

        flight_label = QLabel("flight number:")
        flight_label.setObjectName("normal")
        self.flight_input = QLineEdit()
        # self.flight_input.setMaximumWidth(800)
        # self.flight_input.setMinimumWidth(300)
        self.flight_input.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        carrier_label = QLabel("carrier:")
        carrier_label.setObjectName("normal")
        self.carrier_input = QLineEdit()
        # self.carrier_input.setMaximumWidth(800)
        # self.carrier_input.setMinimumWidth(300)
        self.carrier_input.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        add_fc_button = QPushButton("Add flight+carrier")
        add_fc_button.setFont(QFont("Microsoft YaHei", 12))
        # add_fc_button.setMaximumWidth(800)
        add_fc_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        add_fc_button.setStyleSheet("padding: 10px;")
        add_fc_button.clicked.connect(self.add_fc)

        flight_carrier_layout.addWidget(flight_label)
        flight_carrier_layout.addWidget(self.flight_input)
        flight_carrier_layout.addWidget(carrier_label)
        flight_carrier_layout.addWidget(self.carrier_input)
        flight_carrier_layout.addSpacing(10)
        flight_carrier_layout.addWidget(add_fc_button)

        flight_carrier_layout.setAlignment(Qt.AlignHCenter)  # 整个控件居中
        self.inner_layout.addLayout(flight_carrier_layout)
        self.inner_layout.addSpacing(10)

    def build_fc_manifest_panel(self):

        fc_manifest_layout = QHBoxLayout()

        # 左侧 flight+carrier 列表
        left_layout = QVBoxLayout()
        flight_carrier_label = QLabel("flight+carrier:")
        flight_carrier_label.setObjectName("normal")

        self.fc_list = QTableWidget()
        self.fc_list.setColumnCount(2)
        self.fc_list.setHorizontalHeaderLabels(["Flight", "Carrier"])

        self.fc_list.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.fc_list.setMinimumHeight(300)
        # self.fc_list.setMaximumWidth(300)
        # self.fc_list.setMinimumWidth(200)
        self.fc_list.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.fc_list.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.fc_list.horizontalHeader().setStretchLastSection(True)
        self.fc_list.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.fc_list.verticalHeader().setVisible(False)
        self.fc_list.setSelectionMode(QAbstractItemView.SingleSelection)  # 单选模式

        self.fc_list.setStyleSheet("""
            QTableWidget {
                background-color: #ffffff;
                gridline-color: #ccc;
                border: 1px solid #aaa;
            }
            QHeaderView::section {
                background-color: #0078d7;
                color: white;
                padding: 6px 12px;
                border: 1px solid #aaa;
            }
            QTableWidget::item:selected {
                background-color: #87cefa;
                color: black;
            }
        """)
        self.fc_list.itemClicked.connect(self.show_manifests)

        # TODO:按钮设置样式
        fc_btn_layout = QHBoxLayout()

        table_select_all_btn = QPushButton("Select All / Unselect All")
        table_select_all_btn.setObjectName("fcAddButton")
        table_select_all_btn.clicked.connect(self.table_toggle_selects)

        del_fc_btn = QPushButton("Delete Selected flight+carrier")
        del_fc_btn.setObjectName("fcDelButton")
        del_fc_btn.clicked.connect(self.delete_selected_fc)

        fc_btn_layout.addWidget(table_select_all_btn)
        fc_btn_layout.addWidget(del_fc_btn)

        left_layout.addWidget(flight_carrier_label)
        left_layout.addWidget(self.fc_list)
        left_layout.addLayout(fc_btn_layout)

        # 右侧 manifest 文件列表
        right_layout = QVBoxLayout()
        manifest_label = QLabel("manifest file:")
        manifest_label.setObjectName("normal")
        self.manifest_list = QListWidget()
        self.manifest_list.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.manifest_list.setMinimumHeight(300)
        # self.manifest_list.setMaximumWidth(700)
        # self.manifest_list.setMinimumWidth(550)

        # 支持多选
        self.manifest_list.setSelectionMode(QAbstractItemView.NoSelection)

        self.manifest_list.setStyleSheet("""
            QListWidget {
                background-color: #f9f9f9;
                border: 1px solid #ccc;
                padding: 5px;
            }
            QListWidget::item {
                padding: 5px;
            }
            QListWidget::item:selected {
                background-color: #87cefa;
                color: black;
            }
        """)

        # manifest 操作按钮
        manifest_btn_layout = QHBoxLayout()

        select_all_btn = QPushButton("Select All / Unselect All")
        select_all_btn.setObjectName("fcAddButton")
        select_all_btn.clicked.connect(self.toggle_selects)

        add_manifest_btn = QPushButton("Add Manifest")
        add_manifest_btn.clicked.connect(self.add_manifest)

        del_manifest_btn = QPushButton("Delete Selected Manifest")
        del_manifest_btn.clicked.connect(self.delete_selected_manifest)

        add_manifest_btn.setObjectName("fcAddButton")
        del_manifest_btn.setObjectName("fcDelButton")

        manifest_btn_layout.addWidget(select_all_btn)
        manifest_btn_layout.addWidget(add_manifest_btn)
        manifest_btn_layout.addWidget(del_manifest_btn)

        right_layout.addWidget(manifest_label)
        right_layout.addWidget(self.manifest_list)
        right_layout.addLayout(manifest_btn_layout)

        fc_manifest_group = QFrame()
        fc_manifest_group.setObjectName("fcFrame")
        fc_manifest_group.setFrameShape(QFrame.NoFrame)
        fc_manifest_group.setLayout(fc_manifest_layout)

        # fc_manifest_layout.addStretch()
        fc_manifest_layout.addLayout(left_layout)
        fc_manifest_layout.addSpacing(50)
        fc_manifest_layout.addLayout(right_layout)
        # fc_manifest_layout.addStretch()

        fc_manifest_layout.setStretchFactor(left_layout, 1)
        fc_manifest_layout.setStretchFactor(right_layout, 2)

        # self.inner_layout.addLayout(fc_manifest_layout)
        self.inner_layout.addWidget(fc_manifest_group)

        self.inner_layout.addSpacing(20)

    def build_action_buttons(self):

        save_btn_layout = QHBoxLayout()

        save_email_button = QPushButton("Save Mail Version")
        save_email_button.setFont(QFont("Microsoft YaHei", 12))
        save_email_button.setStyleSheet("padding: 10px;")
        save_email_button.clicked.connect(self.save_mail_version)

        save_system_button = QPushButton("Save System Version")
        save_system_button.setFont(QFont("Microsoft YaHei", 12))
        save_system_button.setStyleSheet("padding: 10px;")
        save_system_button.clicked.connect(self.save_system_version)

        save_btn_layout.addWidget(save_email_button)
        save_btn_layout.addSpacing(20)
        save_btn_layout.addWidget(save_system_button)

        self.inner_layout.addLayout(save_btn_layout)
        self.inner_layout.addSpacing(20)

        # 按钮水平布局（发送按钮和取消按钮并排）
        button_layout = QHBoxLayout()

        # 发送按钮
        submit_button = QPushButton("Submit")
        submit_button.setObjectName("confirmbutton")
        submit_button.clicked.connect(self.submit)

        # 取消按钮
        cancel_button = QPushButton("Cancel")
        cancel_button.setObjectName("cancelButton")
        cancel_button.clicked.connect(self.on_cancel)  # 连接到提示框槽函数

        button_layout.addWidget(submit_button)
        button_layout.addWidget(cancel_button)

        # 将按钮布局添加到主布局中
        self.inner_layout.addLayout(button_layout)
        self.inner_layout.addSpacing(20)

        # clear_button = QPushButton("Clear")
        # clear_button.setFont(QFont("Microsoft YaHei", 12))
        # clear_button.setStyleSheet("padding: 10px;")
        # clear_button.clicked.connect(self.reset)
        # self.inner_layout.addWidget(clear_button)

        # 底部按钮布局
        self.inner_layout.addStretch()

    def add_fc(self):
        flight = self.flight_input.text().strip()
        carrier = self.carrier_input.text().strip()
        if not flight or not carrier:
            QMessageBox.warning(self, "Warning", "Both flight number and carrier must be filled")
            return

        # 避免重复添加
        if any(fc['flight'] == flight and fc['carrier'] == carrier for fc in self.flights_data):
            QMessageBox.warning(self, "Warning", "This flight number+carrier already exists")
            return

        row = self.fc_list.rowCount()
        self.fc_list.insertRow(row)

        # Flight (带checkbox)
        flight_item = QTableWidgetItem(flight)
        flight_item.setFlags(flight_item.flags() | Qt.ItemIsUserCheckable)
        flight_item.setCheckState(Qt.Unchecked)

        self.fc_list.setItem(row, 0, flight_item)
        self.fc_list.setItem(row, 1, QTableWidgetItem(carrier))

        self.flights_data.append({'flight': flight, 'carrier': carrier, 'manifests': []})

        self.flight_input.clear()
        self.carrier_input.clear()

    def show_manifests(self, item):
        self.manifest_list.clear()

        row = self.fc_list.row(item)  # ✅ QTableWidgetItem → int
        if row < 0 or row >= len(self.flights_data):
            return

        for manifest_path in self.flights_data[row]["manifests"]:
            list_item = QListWidgetItem(manifest_path)
            list_item.setToolTip(manifest_path)
            list_item.setFlags(list_item.flags() | Qt.ItemIsUserCheckable)
            list_item.setCheckState(Qt.Unchecked)
            self.manifest_list.addItem(list_item)

    def table_toggle_selects(self):
        if self.fc_list.rowCount() == 0:
            return

        # 判断是否全部勾选
        all_checked = True
        for row in range(self.fc_list.rowCount()):
            item = self.fc_list.item(row, 0)  # 第一列是 checkbox
            if item.checkState() != Qt.Checked:
                all_checked = False
                break

        new_state = Qt.Unchecked if all_checked else Qt.Checked

        for row in range(self.fc_list.rowCount()):
            self.fc_list.item(row, 0).setCheckState(new_state)

    def toggle_selects(self):
        if self.manifest_list.count() == 0:
            return

        # 判断是否已经全部选中
        all_checked = True
        for i in range(self.manifest_list.count()):
            if self.manifest_list.item(i).checkState() != Qt.Checked:
                all_checked = False
                break

        # 如果全部选中 → 取消
        new_state = Qt.Unchecked if all_checked else Qt.Checked

        for i in range(self.manifest_list.count()):
            self.manifest_list.item(i).setCheckState(new_state)

    def delete_selected_fc(self):
        rows_to_delete = []

        for row in range(self.fc_list.rowCount()):
            item = self.fc_list.item(row, 0)

            if item.checkState() == Qt.Checked:
                rows_to_delete.append(row)

        if not rows_to_delete:
            QMessageBox.warning(self, "Warning", "Please select flight+carrier to delete")
            return

        for row in reversed(rows_to_delete):
            self.fc_list.removeRow(row)
            self.flights_data.pop(row)

        self.manifest_list.clear()  # 同时清空右侧 manifest

    def add_manifest(self):
        selected_fc_items = self.fc_list.selectedItems()
        if not selected_fc_items:
            QMessageBox.warning(self, "Warning", "Please select a flight number+carrier first")
            return

        file_dialog = QFileDialog(self)
        file_dialog.setWindowTitle('Select Manifest Files')
        file_dialog.setFileMode(QFileDialog.ExistingFiles)  # 只允许选择一个文件
        file_dialog.setNameFilter("Excel Files (*.xls *.xlsx)")

        if file_dialog.exec_():
            fc_idx = self.fc_list.currentRow()
            for file_path in file_dialog.selectedFiles():
                if file_path in self.flights_data[fc_idx]['manifests']:
                    continue
                self.flights_data[fc_idx]['manifests'].append(file_path)

                list_item = QListWidgetItem(file_path)
                list_item.setToolTip(file_path)  # 鼠标悬浮显示完整路径
                list_item.setFlags(list_item.flags() | Qt.ItemIsUserCheckable)  # 可勾选
                list_item.setCheckState(Qt.Unchecked)
                self.manifest_list.addItem(list_item)

    def delete_selected_manifest(self):
        fc_idx = self.fc_list.currentRow()
        if fc_idx < 0:
            return
        # 倒序删除，避免索引错乱
        for row in reversed(range(self.manifest_list.count())):
            item = self.manifest_list.item(row)
            if item.checkState() == Qt.Checked:
                self.manifest_list.takeItem(row)
                self.flights_data[fc_idx]['manifests'].pop(row)

    def select_file(self):

        file_dialog = QFileDialog(self)
        file_dialog.setWindowTitle('Select Manifest File')
        file_dialog.setFileMode(QFileDialog.ExistingFile)  # 只允许选择一个文件
        file_dialog.setNameFilter("Excel Files (*.xls *.xlsx)")

        if file_dialog.exec_():
            selected_file = file_dialog.selectedFiles()[0]  # 只取第一个
            self.file_path = selected_file  # 保存文件路径
            self.label_status.setText(self.file_path)  # 显示文件路径

    def get_data(self):
        # 返回存储的输入数据
        self.flight = self.input_flight.text()
        self.carrier = self.input_carrier.text()

    def store_data(self, flight, carrier, file_path):
        global_state.flight = flight
        global_state.carrier = carrier
        global_state.manifest_file_path = file_path

    def reset(self):
        # 清除输入框中的文本
        self.input_flight.clear()
        self.input_carrier.clear()

        # 重置文件路径和上传按钮文本
        self.file_path = ""
        self.label_status.setText("No file selected")

    def submit(self):
        # 禁用按钮，改样式、文字
        self.findChild(QPushButton, "confirmbutton").setEnabled(False)
        self.findChild(QPushButton, "confirmbutton").setText("Processing...")

        QApplication.processEvents()

        try:
            self.get_data()
            if not self.flight or not self.carrier or not self.file_path:
                QMessageBox.warning(self, "Missing Info", "Please fill all fields and upload manifest file.")
                return

            request_response = self.api.add_awb(self.flight, self.carrier, self.file_path)

            # self.finished = True
            # self.store_data(self.flight, self.carrier, self.file_path)
            # self.finished_successful.emit()
            # TODO: 全部转换成这种形式
            try:
                if request_response.status_code == 200:
                    response_json = request_response.json()
                    print("get response successfully.")
                    if response_json.get('success', False):
                        print("Add AWB successfully.")
                        QMessageBox.information(self, "Success",
                                                "Add AWB successfully! Please save the process document.")
                        self.store_data(self.flight, self.carrier, self.file_path)
                    else:
                        err_msg = response_json.get('error', 'Unknown error')
                        print(f"Add AWB failed: {err_msg}")
                        QMessageBox.critical(self, "Error", f"Add AWB failed! {err_msg}")
                else:
                    print(f"Failed to add AWB. Status code: {request_response.status_code}")
                    print("Raw response text:", request_response.text)
                    QMessageBox.critical(self, "Error",
                                         f"Failed to add AWB. Status code: {request_response.status_code}")
            except AttributeError:
                print("Error: Response object does not have a status_code attribute.")
                QMessageBox.critical(self, "Error", "Invalid response object received.")

        finally:
            # 恢复按钮状态
            self.findChild(QPushButton, "confirmbutton").setEnabled(True)
            self.findChild(QPushButton, "confirmbutton").setText("Submit")
            QApplication.processEvents()

    def on_cancel(self):
        # 创建提示框，确认取消操作
        reply = QMessageBox.question(self, 'Confirm Cancel',
                                     "This action will discard all changes and return to the first page. Are you sure you want to continue?",
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

        if reply == QMessageBox.Yes:
            self.cancel_requested.emit()  # 发送信号，通知父窗口取消操作
        else:
            print("User canceled the action.")

    def save_mail_version(self):

        if not self.flights_data:
            QMessageBox.warning(self, "No Data", "Please add at least one flight number+carrier and manifest file.")
            return

        # save_dir = QFileDialog.getExistingDirectory(self, "Select Folder to Save Files")
        # if not save_dir:
        #     return  # 用户取消

        for flight_entry in self.flights_data:
            carrier = flight_entry['carrier']
            flight = flight_entry['flight']
            manifests = flight_entry.get('manifests', [])

            for manifest_path in manifests:
                try:
                    # df = pd.read_excel(manifest_path, dtype=str)

                    # === 使用 openpyxl 精准读取，避免精度与前导0丢失 ===
                    wb = load_workbook(manifest_path, data_only=True)
                    ws = wb.active
                    data = []
                    for row in ws.iter_rows(values_only=True):
                        # === 新增：跳过整行为空（None或空字符串）的行 ===
                        if not any(cell is not None and str(cell).strip() != '' for cell in row):
                            continue
                        data.append(list(row))
                    df = pd.DataFrame(data[1:], columns=data[0])
                    df = df.fillna('')  # 避免 NaN 干扰

                    normalized_cols = [str(col).replace(" ", "").replace("\n", "").lower() for col in df.columns]

                except Exception as e:
                    QMessageBox.critical(self, "Read Error", f"Failed to read Excel file:\n{manifest_path}\n{e}")
                    return

                if not any("airwaybill" in col for col in normalized_cols):
                    QMessageBox.warning(self, "Missing Column",
                                        f"Column 'AirWayBill' is not found in Excel:\n{manifest_path}.")
                    return

                # 提取原始列名
                # airwaybill_col = df.columns[
                #     normalized_cols.index(next(col for col in normalized_cols if "airwaybill" in col))]

                # 提取唯一值
                # airway_bills = df[airwaybill_col].dropna().astype(str).unique()
                #
                # if len(airway_bills) > 1:
                #     QMessageBox.warning(self, "Multiple AirWayBills",
                #                         f"Only one unique AirWayBill is supported. Found: {', '.join(airway_bills)} in Excel:\n{manifest_path}.")
                #     return

                if not any("boxnumber" in col for col in normalized_cols):
                    QMessageBox.warning(self, "Missing Column",
                                        f"Column 'BOXNumber' is not found in Excel:\n{manifest_path}.")
                    return

                # 提取原始列名
                box_col = df.columns[normalized_cols.index(next(col for col in normalized_cols if "boxnumber" in col))]

                if not any("trackingnumber" in col or "parcelnumber" in col for col in normalized_cols):
                    QMessageBox.warning(self, "Missing Column",
                                        f"Column 'TrackingNumber' or 'ParcelNumber' is not found in Excel:\n{manifest_path}.")
                    return

                # 提取原始列名
                tc_col = df.columns[normalized_cols.index(
                    next(col for col in normalized_cols if "trackingnumber" in col or "parcelnumber" in col))]

                # box_df = df[required_columns].drop_duplicates()
                # box_df = box_df.sort_values(by='BOXNumber')

                # 提取并去重
                # box_df = df[[box_col, 'TrackingNumber']].drop_duplicates()
                # box_df = box_df.sort_values(by=box_col)

                box_df = df[[box_col, tc_col]].copy()
                box_df['orig_index'] = df.index  # 保存原始行索引
                box_df = box_df.sort_values(by=box_col, kind='stable')

                # 统一改名
                box_df.columns = ['BOXNumber', 'TrackingNumber', 'orig_index']

                # 置空重复的 BOX NUMBER 以形成视觉缩进
                # box_df['BOXNumber'] = box_df['BOXNumber'].mask(box_df['BOXNumber'].duplicated())

                # === 新增的四个 MRN 列 ===
                mrn_cols = [
                    'MRN(ARN) from ICS2\n EU based on AWB - F21',
                    'MRN(ENS) from ICS2 EU\n based on each box - F26',
                    'MRN(CLEARANCES) from\n ICS2 EU based on Arrival\n Notification',
                    'MRN(RELEASES) from\n ZC429HUB'
                ]
                for col in mrn_cols:
                    box_df[col] = ''

                box_df['FlightNumber'] = flight

                # === 把原表中剩下的列拼回去（保持原顺序） ===
                extra_cols = [c for c in df.columns if c not in [box_col, tc_col] and c not in [None, '']]
                for col in extra_cols:
                    # 使用 loc 通过原始索引赋值，保证对齐
                    box_df[col] = df.loc[box_df['orig_index'], col].values

                # === 最后新增 Dsk 列 ===
                box_df['Dsk'] = ''

                # 最后删除 orig_index 列
                box_df = box_df.drop(columns=['orig_index'])

                # manifest_name = os.path.splitext(os.path.basename(manifest_path))[0]
                # file_name = f"{flight}_{carrier}_{manifest_name}_process_document.xlsx"
                default_dir = os.path.dirname(manifest_path)

                file_name = f"{flight}_{carrier}_process_document_mail_version.xlsx"
                save_path, _ = QFileDialog.getSaveFileName(
                    self,
                    "Save Process Document",
                    os.path.join(default_dir, file_name),  # 默认路径 + 默认文件名
                    "Excel Files (*.xls *.xlsx)"
                )
                # save_path = os.path.join(save_dir, file_name)
                if not save_path:
                    return
                try:
                    box_df.to_excel(save_path, engine='openpyxl', index=False)

                    # 打开文件，设置列宽
                    wb = openpyxl.load_workbook(save_path)
                    ws = wb.active

                    # 设置不加粗标题（可选）
                    for cell in ws[1]:
                        cell.font = openpyxl.styles.Font(bold=False)

                    # 设置所有列列宽为固定值，比如 25
                    for col_idx in range(1, ws.max_column + 1):
                        col_letter = get_column_letter(col_idx)
                        ws.column_dimensions[col_letter].width = 25

                    # === 所有单元格居中（表头 + 正文） ===
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                        for cell in row:
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                    # === 新增：MRN 四列表头标黄 ===
                    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    for col_idx in range(3, 7):  # MRN 四列对应第3到第6列
                        ws.cell(row=1, column=col_idx).fill = header_fill

                    # === 表头行高（让换行显示完整） ===
                    ws.row_dimensions[1].height = 40

                    # === 正文行高（比如每行 25，可按需要调整） ===
                    for row_idx in range(2, ws.max_row + 1):
                        ws.row_dimensions[row_idx].height = 25

                    # 新增：给相同 BOXNumber 的第一条标红
                    box_number_col = 1  # BOXNumber 在第1列
                    # seen = set()
                    # for row in range(2, ws.max_row + 1):  # 从第二行开始（跳过表头）
                    #     cell = ws.cell(row=row, column=box_number_col)
                    #     value = cell.value
                    #     if value and value not in seen:
                    #         cell.font = Font(color="FF0000")  # 红色字体
                    #         seen.add(value)
                    current_value = None
                    for row in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row, column=box_number_col)
                        value = cell.value
                        if value != current_value:
                            # 新的一组，首条红色
                            cell.font = Font(color="FF0000")
                            current_value = value

                    # 保存修改
                    wb.save(save_path)

                    print(f"Saved processed file: {save_path}")

                except Exception as e:
                    QMessageBox.critical(self, "Save Error", f"Failed to save Process file:\n{manifest_path}\n{e}")

        QMessageBox.information(self, "Success", f"Process Files successfully saved to:\n{save_path}")

    def save_system_version(self):
        if not self.flights_data:
            QMessageBox.warning(self, "No Data", "Please add at least one flight number+carrier and manifest file.")
            return

        # save_dir = QFileDialog.getExistingDirectory(self, "Select Folder to Save Files")
        # if not save_dir:
        #     return  # 用户取消

        sum_cols_keywords = ["totalprice", "grossmasskg", "netmasskg", "quantity", "amountpackages"]

        for flight_entry in self.flights_data:
            carrier = flight_entry['carrier']
            flight = flight_entry['flight']
            manifests = flight_entry.get('manifests', [])

            for manifest_path in manifests:
                try:
                    # df = pd.read_excel(manifest_path, dtype=str)

                    # === 使用 openpyxl 精准读取，避免精度与前导0丢失 ===
                    wb = load_workbook(manifest_path, data_only=True)
                    ws = wb.active
                    data = []
                    for row in ws.iter_rows(values_only=True):
                        # === 新增：跳过整行为空（None或空字符串）的行 ===
                        if not any(cell is not None and str(cell).strip() != '' for cell in row):
                            continue
                        data.append(list(row))
                    df = pd.DataFrame(data[1:], columns=data[0])
                    df = df.fillna('')  # 避免 NaN 干扰

                    normalized_cols = [str(col).replace(" ", "").replace("\n", "").lower() for col in df.columns]

                except Exception as e:
                    QMessageBox.critical(self, "Read Error", f"Failed to read Excel file:\n{manifest_path}\n{e}")
                    return

                if not any("airwaybill" in col for col in normalized_cols):
                    QMessageBox.warning(self, "Missing Column",
                                        f"Column 'AirWayBill' is not found in Excel:\n{manifest_path}.")
                    return

                # # 提取原始列名
                # airwaybill_col = df.columns[
                #     normalized_cols.index(next(col for col in normalized_cols if "airwaybill" in col))]

                # 提取唯一值
                # airway_bills = df[airwaybill_col].dropna().astype(str).unique()
                #
                # if len(airway_bills) > 1:
                #     QMessageBox.warning(self, "Multiple AirWayBills",
                #                         f"Only one unique AirWayBill is supported. Found: {', '.join(airway_bills)} in Excel:\n{manifest_path}.")
                #     return

                if not any("boxnumber" in col for col in normalized_cols):
                    QMessageBox.warning(self, "Missing Column",
                                        f"Column 'BOXNumber' is not found in Excel:\n{manifest_path}.")
                    return

                # 提取原始列名
                box_col = df.columns[
                    normalized_cols.index(next(col for col in normalized_cols if "boxnumber" in col))]

                if not any("trackingnumber" in col or "parcelnumber" in col for col in normalized_cols):
                    QMessageBox.warning(self, "Missing Column",
                                        f"Column 'TrackingNumber' or 'ParcelNumber' is not found in Excel:\n{manifest_path}.")
                    return

                # 提取原始列名
                tc_col = df.columns[normalized_cols.index(
                    next(col for col in normalized_cols if "trackingnumber" in col or "parcelnumber" in col))]

                # box_df = df[required_columns].drop_duplicates()
                # box_df = box_df.sort_values(by='BOXNumber')

                # 提取并去重
                # box_df = df[[box_col, 'TrackingNumber']].drop_duplicates()
                # box_df = box_df.sort_values(by=box_col)

                box_df = df[[box_col, tc_col]].copy()
                box_df['orig_index'] = df.index  # 保存原始行索引
                box_df = box_df.sort_values(by=box_col, kind='stable')

                # 统一改名
                box_df.columns = ['BOXNumber', 'TrackingNumber', 'orig_index']

                # 置空重复的 BOX NUMBER 以形成视觉缩进
                # box_df['BOXNumber'] = box_df['BOXNumber'].mask(box_df['BOXNumber'].duplicated())

                # === 新增的四个 MRN 列 ===
                mrn_cols = [
                    'MRN(ARN) from ICS2\n EU based on AWB - F21',
                    'MRN(ENS) from ICS2 EU\n based on each box - F26',
                    'MRN(CLEARANCES) from\n ICS2 EU based on Arrival\n Notification',
                    'MRN(RELEASES) from\n ZC429HUB'
                ]

                # === 把原表中剩下的列拼回去（保持原顺序） ===
                df['TrackingNumber'] = df[tc_col]
                df['BOXNumber'] = df[box_col]

                agg_dict = {}

                for col in df.columns:
                    if col in [tc_col, box_col]:
                        continue
                    col_norm = str(col).replace(" ", "").replace("\n", "").lower()
                    if any(k in col_norm for k in sum_cols_keywords):
                        agg_dict[col] = 'sum'
                        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
                    else:
                        agg_dict[col] = 'first'

                # BOXNumber 列保留第一个出现值
                agg_dict['BOXNumber'] = 'first'

                merged_df = df.groupby(['TrackingNumber'], as_index=False).agg(agg_dict)

                for col in mrn_cols:
                    merged_df[col] = ''
                merged_df['FlightNumber'] = flight
                # === 最后新增 Dsk 列 ===
                merged_df['Dsk'] = ''

                if 'orig_index' in merged_df.columns:
                    merged_df = merged_df.drop(columns=['orig_index'])

                other_cols = [c for c in merged_df.columns if
                              c not in ['BOXNumber', 'TrackingNumber'] + mrn_cols + ['FlightNumber', 'Dsk']]
                final_df = merged_df[
                    ['BOXNumber', 'TrackingNumber'] + mrn_cols + ['FlightNumber'] + other_cols + ['Dsk']]

                # 按 BOXNumber 排序
                final_df = final_df.sort_values(by='BOXNumber', kind='stable').reset_index(drop=True)

                # manifest_name = os.path.splitext(os.path.basename(manifest_path))[0]
                # file_name = f"{flight}_{carrier}_{manifest_name}_process_document.xlsx"
                default_dir = os.path.dirname(manifest_path)

                file_name = f"{flight}_{carrier}_process_document_system_version.xlsx"
                save_path, _ = QFileDialog.getSaveFileName(
                    self,
                    "Save Process Document",
                    os.path.join(default_dir, file_name),  # 默认路径 + 默认文件名
                    "Excel Files (*.xls *.xlsx)"
                )
                # save_path = os.path.join(save_dir, file_name)
                if not save_path:
                    return

                try:
                    final_df.to_excel(save_path, engine='openpyxl', index=False)

                    # 打开文件，设置列宽
                    wb = openpyxl.load_workbook(save_path)
                    ws = wb.active

                    # 设置不加粗标题（可选）
                    for cell in ws[1]:
                        cell.font = openpyxl.styles.Font(bold=False)

                    # 设置所有列列宽为固定值，比如 25
                    for col_idx in range(1, ws.max_column + 1):
                        col_letter = get_column_letter(col_idx)
                        ws.column_dimensions[col_letter].width = 25

                    # === 所有单元格居中（表头 + 正文） ===
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                        for cell in row:
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                    # === 新增：MRN 四列表头标黄 ===
                    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    for col_idx in range(3, 7):  # MRN 四列对应第3到第6列
                        ws.cell(row=1, column=col_idx).fill = header_fill

                    # === 表头行高（让换行显示完整） ===
                    ws.row_dimensions[1].height = 40

                    # === 正文行高（比如每行 25，可按需要调整） ===
                    for row_idx in range(2, ws.max_row + 1):
                        ws.row_dimensions[row_idx].height = 25

                    # 新增：给相同 BOXNumber 的第一条标红
                    box_number_col = 1  # BOXNumber 在第1列
                    # seen = set()
                    # for row in range(2, ws.max_row + 1):  # 从第二行开始（跳过表头）
                    #     cell = ws.cell(row=row, column=box_number_col)
                    #     value = cell.value
                    #     if value and value not in seen:
                    #         cell.font = Font(color="FF0000")  # 红色字体
                    #         seen.add(value)
                    current_value = None
                    for row in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row, column=box_number_col)
                        value = cell.value
                        if value != current_value:
                            # 新的一组，首条红色
                            cell.font = Font(color="FF0000")
                            current_value = value

                    for col_idx in range(1, ws.max_column + 1):
                        col_letter = get_column_letter(col_idx)
                        for cell in ws[col_letter]:
                            if isinstance(cell.value, float):
                                cell.number_format = 'General'  # Excel 自动显示合理的小数位

                    # 保存修改
                    wb.save(save_path)

                    print(f"Saved processed file: {save_path}")

                except Exception as e:
                    QMessageBox.critical(self, "Save Error", f"Failed to save Process file:\n{manifest_path}\n{e}")

        QMessageBox.information(self, "Success", f"Process Files successfully saved to:\n{save_path}")
