import pandas as pd
from openpyxl.reader.excel import load_workbook

# === 用户配置路径 ===
box_file = "871-54579254_E-commercet.xlsx"  # 包含 BOXNumber 和 TrackingNumber 的文件
mapping_file = "871-54579254_1_process_document.xlsx"  # 包含 Container code, Tracking code, F26 MRN 的文件
output_file = "871-54579254_E-commercet_Clearance.xlsx"

# === 读取映射表 ===
mapping_df = pd.read_excel(mapping_file, dtype=str).fillna('')
# mapping_dict = {(row['BOXNumber'], row['TrackingNumber']): row["""MRN(RELEASES) from
#  ZC429HUB"""]
#                 for _, row in mapping_df.iterrows()}

mapping_dict = {(row['BOXNumber'], row['TrackingNumber']): row["""MRN(RELEASES)"""]
                for _, row in mapping_df.iterrows()}

# mapping_dict = {(row['Container code'], row['Tracking code']): row['F26 MRN']
#                 for _, row in mapping_df.iterrows()}
#
# mapping_dict = {(row['Hawb Number'], row['Item No.']): row['Dsk']
#                 for _, row in mapping_df.iterrows()}

# === 打开原始 Excel ===
wb = load_workbook(box_file)
ws = wb.active

# 找到列索引
header = [cell.value for cell in ws[1]]
try:
    box_idx = header.index('Hawb Number') + 1
#     box_idx = header.index("""Box Number/
# Numer Carton""") + 1
#     box_idx = header.index('BOXNumber') + 1
#     box_idx = header.index('Box Number') + 1
    tracking_idx = header.index('Item No.') + 1
#     tracking_idx = header.index("""Parcel Number/
# Numer paczki""") + 1
#     tracking_idx = header.index('TrackingNumber') + 1
    # f26_idx = header.index('MRN(ENS) from ICS2 EU\n based on each box - F26') + 1
    # f26_idx = header.index('Dsk') + 1
    f26_idx = header.index('Clearance') + 1
except ValueError as e:
    raise ValueError(f"列未找到: {e}")

# === 填充 F26 MRN ===
for row in range(2, ws.max_row + 1):
    box_val = ws.cell(row=row, column=box_idx).value
    tracking_val = ws.cell(row=row, column=tracking_idx).value
    key = (str(box_val), str(tracking_val))

    cell = ws.cell(row=row, column=f26_idx)
    old_value = cell.value

    if not old_value:
        new_value = mapping_dict.get(key)
        if new_value:  # ✅ 映射表里也必须真的有值
            cell.value = new_value

# === 保存结果 ===
wb.save(output_file)
print(f"Done! Output saved to {output_file}")